#!/usr/bin/env python3
"""Valida los tres XLSX fuente antes de reconstruir el dashboard."""

from __future__ import annotations

import argparse
from collections import Counter
import json
from pathlib import Path

from openpyxl import load_workbook

try:
    from .build_dashboard import (
        DEFAULT_CMS, DEFAULT_DIRECTORY, DEFAULT_RESPONSES, DEFAULT_SETTINGS,
        STABILITY_CONTROLS, build_payload, clean_text, file_sha256, find_directory_header,
        find_header, included_store_statuses, is_all_regions, is_no, is_yes, key_text, load_cms, load_settings,
        directory_sheet,
        normalize_ceco, normalize_dm, parse_date, validate_xlsx,
    )
except ImportError:  # Ejecución directa: python scripts/validate_sources.py
    from build_dashboard import (
        DEFAULT_CMS, DEFAULT_DIRECTORY, DEFAULT_RESPONSES, DEFAULT_SETTINGS,
        STABILITY_CONTROLS, build_payload, clean_text, file_sha256, find_directory_header,
        find_header, included_store_statuses, is_all_regions, is_no, is_yes, key_text, load_cms, load_settings,
        directory_sheet,
        normalize_ceco, normalize_dm, parse_date, validate_xlsx,
    )

CMS_SHEETS = {"Actividades", "Gerentes", "Configuracion", "Tiendas Abiertas"}
CMS_CONFIG_KEYS = {
    "projectName", "region", "directorySheet", "onlyOpenStores", "includedStoreStatuses",
    "requireEvidence", "publishEvidenceLinks", "publishPersonalData",
    "evidenceAllowedHosts", "regionalDirectorName", "regionalDirectorPhoto",
}


def duplicates(values: list[str]) -> list[str]:
    return sorted(value for value, count in Counter(values).items() if count > 1)


def validate_cms_engine(path: Path) -> dict[str, int]:
    """Audita encabezados y filas activas sin bloquear borradores ni celdas auxiliares."""
    validate_xlsx(path, "el CMS maestro")
    workbook = load_workbook(path, read_only=False, data_only=False)
    missing_sheets = CMS_SHEETS.difference(workbook.sheetnames)
    if missing_sheets:
        raise ValueError("Faltan hojas CMS: " + ", ".join(sorted(missing_sheets)))

    ws = workbook["Actividades"]
    header_row, cols = find_header(ws, {
        "orden", "actividad", "descripcion", "fecha inicio", "fecha limite",
        "activo", "evidencia requerida", "prioridad", "estado fecha",
    })
    names: list[str] = []
    orders: list[str] = []
    activity_rows = 0
    drafts = 0
    fallback_orders = 0
    manual_statuses = 0
    custom_priorities = 0
    safe_defaults = 0
    invalid_dates = 0
    corrected_date_ranges = 0
    for row_number in range(header_row + 1, ws.max_row + 1):
        name = clean_text(ws.cell(row_number, cols["actividad"] + 1).value)
        if not name:
            continue
        active_value = ws.cell(row_number, cols["activo"] + 1).value
        if not is_yes(active_value):
            drafts += 1
            continue
        activity_rows += 1
        names.append(key_text(name))
        order = clean_text(ws.cell(row_number, cols["orden"] + 1).value)
        try:
            orders.append(str(int(float(order))))
        except (TypeError, ValueError):
            fallback_orders += 1
        evidence_value = ws.cell(row_number, cols["evidencia requerida"] + 1).value
        if not (is_yes(evidence_value) or is_no(evidence_value)):
            safe_defaults += 1
        priority = key_text(ws.cell(row_number, cols["prioridad"] + 1).value)
        if priority and priority not in {"alta", "media", "baja"}:
            custom_priorities += 1
        parsed_dates = []
        for field in ("fecha inicio", "fecha limite"):
            value = ws.cell(row_number, cols[field] + 1).value
            try:
                parsed_dates.append(parse_date(value))
            except ValueError:
                parsed_dates.append(None)
                invalid_dates += 1
        if parsed_dates[0] and parsed_dates[1] and parsed_dates[1] < parsed_dates[0]:
            corrected_date_ranges += 1
        formula = ws.cell(row_number, cols["estado fecha"] + 1).value
        if not isinstance(formula, str) or not formula.startswith("="):
            # Python recalcula el estado por fechas; esta celda es informativa.
            manual_statuses += 1
    if duplicates(names):
        raise ValueError("Actividades CMS duplicadas: " + ", ".join(duplicates(names)))
    if duplicates(orders):
        raise ValueError("Órdenes CMS duplicados: " + ", ".join(duplicates(orders)))

    manager_ws = workbook["Gerentes"]
    manager_header, manager_cols = find_header(manager_ws, {"dm", "nombre corto", "foto webp", "activo"})
    managers: list[str] = []
    for row_number in range(manager_header + 1, manager_ws.max_row + 1):
        dm = clean_text(manager_ws.cell(row_number, manager_cols["dm"] + 1).value)
        if not dm or not is_yes(manager_ws.cell(row_number, manager_cols["activo"] + 1).value):
            continue
        managers.append(key_text(dm))
    if duplicates(managers):
        raise ValueError("Gerentes CMS duplicados: " + ", ".join(duplicates(managers)))

    stores_ws = workbook["Tiendas Abiertas"]
    stores_header, stores_cols = find_header(stores_ws, {"cc", "cc nombre", "region", "estatus", "dm"})
    cms_open_stores = 0
    cms_cecos: list[str] = []
    for row in stores_ws.iter_rows(min_row=stores_header + 1, values_only=True):
        ceco = normalize_ceco(row[stores_cols["cc"]])
        if not ceco:
            continue
        if key_text(row[stores_cols["estatus"]]) != "abierta":
            raise ValueError(f"CMS Tiendas Abiertas contiene un estatus no permitido: {ceco}")
        cms_cecos.append(ceco)
        cms_open_stores += 1
    if duplicates(cms_cecos):
        raise ValueError("CeCo duplicados en CMS Tiendas Abiertas: " + ", ".join(duplicates(cms_cecos)))

    config_ws = workbook["Configuracion"]
    config_header, config_cols = find_header(config_ws, {"clave", "valor"})
    config_keys: list[str] = []
    config_defaults = 0
    boolean_keys = {"onlyOpenStores", "requireEvidence", "publishEvidenceLinks", "publishPersonalData"}
    for row in range(config_header + 1, config_ws.max_row + 1):
        key = clean_text(config_ws.cell(row, config_cols["clave"] + 1).value)
        if not key:
            continue
        config_keys.append(key)
        value = config_ws.cell(row, config_cols["valor"] + 1).value
        if not clean_text(value) or (key in boolean_keys and not (is_yes(value) or is_no(value))):
            config_defaults += 1
    missing_keys = CMS_CONFIG_KEYS.difference(config_keys)
    if duplicates(config_keys):
        raise ValueError("Claves CMS duplicadas: " + ", ".join(duplicates(config_keys)))
    return {
        "activities": activity_rows,
        "managers": len(managers),
        "openStores": cms_open_stores,
        "settings": len(config_keys),
        "drafts": drafts,
        "fallbackOrders": fallback_orders,
        "manualStatuses": manual_statuses,
        "customPriorities": custom_priorities,
        "invalidDates": invalid_dates,
        "correctedDateRanges": corrected_date_ranges,
        "safeDefaults": safe_defaults + config_defaults + len(missing_keys),
    }


def validate_directory_engine(path: Path, cms_path: Path, settings_path: Path) -> dict[str, int | str]:
    """Audita únicamente la hoja operativa configurada; las históricas no alimentan el tablero."""
    _, _, cms_settings, _ = load_cms(cms_path)
    settings = load_settings(settings_path, cms_settings)
    validate_xlsx(path, "el directorio")
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws, header_row, headers = directory_sheet(workbook, settings.get("directorySheet"))
    sheet_name = ws.title
    cols = {key_text(value): index for index, value in enumerate(headers) if clean_text(value)}
    missing_columns = {"cc", "cc nombre", "region", "dm"}.difference(cols)
    if missing_columns:
        raise ValueError("Directorio incompleto: " + ", ".join(sorted(missing_columns)))
    active_cecos: list[str] = []
    missing_dm: list[str] = []
    excluded_stores = 0
    allowed_statuses = included_store_statuses(settings)
    if settings.get("onlyOpenStores") and "estatus" not in cols:
        raise ValueError("El Directorio debe incluir la columna Estatus para publicar sólo tiendas abiertas")
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        ceco = normalize_ceco(row[cols["cc"]])
        if not ceco:
            continue
        region = clean_text(row[cols["region"]])
        status_col = cols.get("estatus")
        status = key_text(row[status_col]) if status_col is not None else ""
        if not is_all_regions(settings.get("region")) and key_text(region) != key_text(settings["region"]):
            continue
        if settings.get("onlyOpenStores") and status not in allowed_statuses:
            excluded_stores += 1
            continue
        active_cecos.append(ceco)
        if normalize_dm(row[cols["dm"]]) == "DM pendiente":
            missing_dm.append(ceco)
    if duplicates(active_cecos):
        raise ValueError("CeCo duplicados en la hoja operativa: " + ", ".join(duplicates(active_cecos)))
    hidden_sheets = sum(sheet.sheet_state != "visible" for sheet in workbook.worksheets)
    regions = {
        clean_text(row[cols["region"]])
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True)
        if normalize_ceco(row[cols["cc"]])
    }
    return {
        "stores": len(active_cecos), "sheet": sheet_name, "hiddenSheets": hidden_sheets,
        "regions": len(regions), "missingDm": len(missing_dm), "excludedStores": excluded_stores,
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Valida integridad, encabezados y cruces de los XLSX fuente.")
    parser.add_argument("--responses", type=Path, default=DEFAULT_RESPONSES)
    parser.add_argument("--directory", type=Path, default=DEFAULT_DIRECTORY)
    parser.add_argument("--cms", type=Path, default=DEFAULT_CMS)
    parser.add_argument("--settings", type=Path, default=DEFAULT_SETTINGS)
    args = parser.parse_args()

    cms_audit = validate_cms_engine(args.cms)
    directory_audit = validate_directory_engine(args.directory, args.cms, args.settings)

    payload = build_payload(args.responses, args.directory, args.settings, args.cms)
    quality = payload["quality"]
    schema = quality["responseSchema"]
    conflicting_evidence = {
        key: rows for key, rows in schema.get("evidenceIssues", {}).items()
        if key not in {"generic-evidence-fallback"} and rows
    }
    blocking_issues = {
        "conflictosFilas": schema.get("rowConflicts", []),
        "conflictosEvidencia": conflicting_evidence,
        "conflictosAplicabilidad": schema.get("applicabilityIssues", {}),
        "cecosDesconocidos": quality.get("unknownCeCos", []),
        "vinculosInseguros": quality.get("unsafeEvidenceRows", []),
    }
    if any(blocking_issues.values()):
        raise SystemExit("Fuentes rechazadas: " + json.dumps(blocking_issues, ensure_ascii=False))

    print(
        "Fuentes XLSX aprobadas · "
        f"Forms {quality['responsesRead']} filas · "
        f"{payload['summary']['stores']} tiendas · "
        f"{payload['summary']['activities']} actividades · "
        f"encabezado Forms fila {schema['headerRow']}"
    )
    print(
        "Control CMS · "
        f"{len(quality.get('hiddenActivityRows', []))} filas Forms ignoradas · "
        f"{len(quality.get('hiddenActivities', []))} actividades ausentes o inactivas · "
        "sin impacto en avance ni fecha de corte"
    )
    controls = quality.get("stabilityControls", {})
    if tuple(controls) != STABILITY_CONTROLS or not all(controls.values()):
        raise SystemExit("Fuentes rechazadas: controles de estabilidad incompletos")
    print(
        f"Estabilidad {quality.get('stabilityScore')} · encabezados dinámicos · filas dinámicas · "
        f"{quality.get('duplicateValidResponses', 0)} respuestas históricas deduplicadas · "
        f"{len(quality.get('canonicalizedActivityRows', []))} nombres similares normalizados"
    )
    print(
        "Motores auditados · "
        f"CMS {cms_audit['activities']} actividades / {cms_audit['managers']} DM / {cms_audit['openStores']} tiendas abiertas / {cms_audit['settings']} parámetros · "
        f"Directorio {directory_audit['stores']} tiendas / {directory_audit['regions']} regiones en {directory_audit['sheet']} · "
        f"{directory_audit['missingDm']} asignaciones DM pendientes · "
        f"{directory_audit['excludedStores']} tiendas fuera por Estatus · "
        f"{directory_audit['hiddenSheets']} hoja histórica fuera del cálculo"
    )
    print(
        "Tolerancia CMS · "
        f"{cms_audit['drafts']} borradores ignorados · "
        f"{cms_audit['fallbackOrders']} órdenes con respaldo · "
        f"{cms_audit['manualStatuses']} estados informativos · "
        f"{cms_audit['customPriorities']} prioridades personalizadas · "
        f"{cms_audit['invalidDates']} fechas con respaldo · "
        f"{cms_audit['correctedDateRanges']} rangos corregidos · "
        f"{cms_audit['safeDefaults']} valores protegidos por defecto"
    )
    for label, path in (("Forms", args.responses), ("Directorio", args.directory), ("CMS", args.cms)):
        print(f"{label}: {path.name} · SHA256 {file_sha256(path)[:12]}")


if __name__ == "__main__":
    main()
