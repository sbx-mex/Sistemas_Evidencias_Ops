#!/usr/bin/env python3
"""Construye el dashboard estático desde Forms + Directorio multirregión.

Fuentes editoriales:
  - cms/Sistema de Evidencias OPS.xlsx
  - cms/Directorio.xlsx
  - cms/Sistema_Evidencias_OPS_CMS.xlsx

El navegador nunca procesa los Excel. Este motor valida encabezados, cruza CeCo,
deduplica cumplimiento por tienda/actividad y publica únicamente el JSON mínimo.
"""

from __future__ import annotations

import argparse
from difflib import SequenceMatcher
import hashlib
import json
import re
import unicodedata
import zipfile
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlsplit

from openpyxl import load_workbook

try:
    from .io_utils import atomic_write_text
except ImportError:  # Ejecución directa: python scripts/build_dashboard.py
    from io_utils import atomic_write_text

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_RESPONSES = ROOT / "cms" / "Sistema de Evidencias OPS.xlsx"
DEFAULT_DIRECTORY = ROOT / "cms" / "Directorio.xlsx"
DEFAULT_CMS = ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx"
DEFAULT_SETTINGS = ROOT / "config" / "settings.json"
DEFAULT_OUTPUT = ROOT / "data" / "dashboard.json"

RESPONSE_FIELDS = {
    "id": ("Id",),
    "started": ("Hora de inicio",),
    "finished": ("Hora de finalización", "Hora de finalizacion"),
    "email": ("Correo electrónico", "Correo electronico"),
    "name": ("Nombre",),
    "activity": ("Selecciona la actividad que deseas registrar", "Actividad"),
    "ceco": ("CeCo", "CC", "Centro de costo"),
}

CONFIRMATION_HEADERS = (
    "¿Confirmas que realizaste la actividad seleccionada?",
    "Confirmación",
    "Confirmacion",
)
EVIDENCE_HEADERS = ("Evidencia", "Evidencia del avance")
APPLICABILITY_HEADER_HINTS = (
    "aplica", "aplican", "tienes", "tiene", "cuentas con", "cuenta con",
    "dispones", "dispone", "participa", "participan",
)
REQUIRED_RESPONSE_FIELDS = {"activity", "ceco"}
REQUIRED_XLSX_MEMBERS = {"[Content_Types].xml", "xl/workbook.xml", "xl/_rels/workbook.xml.rels"}
MOJIBAKE_MARKERS = ("\u00c3", "\u00c2", "\u00e2")
STABILITY_CONTROLS = (
    "sourceIntegrity",
    "cmsActiveAllowlist",
    "externalFormsIsolation",
    "canonicalActivityMatching",
    "evidenceHeaderSafety",
    "columnOrderIndependence",
    "duplicateResponseResolution",
    "directoryUniqueness",
    "safeEvidenceLinks",
    "atomicPublication",
)
KNOWN_SETTING_KEYS = (
    "projectName", "region", "directorySheet", "onlyOpenStores", "includedStoreStatuses",
    "requireEvidence", "publishEvidenceLinks", "publishPersonalData",
    "evidenceAllowedHosts", "regionalDirectorName", "regionalDirectorPhoto",
)


def repair_mojibake(value: str) -> str:
    """Repara UTF-8 interpretado como Latin-1 sin alterar texto Unicode válido."""
    repaired = value
    for _ in range(2):
        if not any(marker in repaired for marker in MOJIBAKE_MARKERS):
            break
        try:
            candidate = repaired.encode("latin-1").decode("utf-8")
        except (UnicodeEncodeError, UnicodeDecodeError):
            break
        before = sum(repaired.count(marker) for marker in MOJIBAKE_MARKERS)
        after = sum(candidate.count(marker) for marker in MOJIBAKE_MARKERS)
        if after >= before:
            break
        repaired = candidate
    return repaired


def clean_text(value: Any) -> str:
    return re.sub(r"\s+", " ", repair_mojibake(str(value or "").strip()))


def key_text(value: Any) -> str:
    text = unicodedata.normalize("NFD", clean_text(value).casefold())
    return "".join(char for char in text if unicodedata.category(char) != "Mn")


def compact_key(value: Any) -> str:
    """Clave tolerante a espacios, guiones, &, acentos y cambios de mayúsculas."""
    return re.sub(r"[^a-z0-9]+", "", key_text(value))


def normalize_dm(value: Any) -> str:
    dm = clean_text(value)
    if key_text(dm) in {"", "na", "n/a", "sin dm", "cierre de ceco"}:
        return "DM pendiente"
    return dm


def active_activity_catalog(activities: list[dict[str, Any]]) -> tuple[dict[str, str], dict[str, str]]:
    """Crea el catálogo autorizado por CMS y rechaza nombres activos ambiguos.

    La primera clave conserva palabras y la segunda tolera signos, espacios y
    acentos. Esto permite que Forms cambie el formato visual del nombre sin
    convertir una actividad ajena al CMS en un cumplimiento válido.
    """
    by_text: dict[str, str] = {}
    by_compact: dict[str, str] = {}
    for item in activities:
        name = clean_text(item.get("name"))
        text_key = key_text(name)
        compact = compact_key(name)
        if not name or not text_key or not compact:
            raise ValueError("El CMS contiene una actividad activa sin nombre válido")
        if text_key in by_text or compact in by_compact:
            duplicate = by_text.get(text_key) or by_compact.get(compact) or name
            raise ValueError(f"El CMS contiene actividades activas duplicadas: {duplicate} / {name}")
        by_text[text_key] = name
        by_compact[compact] = name
    return by_text, by_compact


def canonical_cms_activity(value: Any, by_text: dict[str, str], by_compact: dict[str, str]) -> str | None:
    """Devuelve exclusivamente una actividad activa CMS con coincidencia única.

    Primero exige igualdad normalizada. Un error menor de escritura sólo se
    acepta con afinidad alta y distancia suficiente frente a la segunda opción.
    Si hay duda, la fila queda fuera del cálculo en vez de adivinar.
    """
    exact = by_text.get(key_text(value)) or by_compact.get(compact_key(value))
    if exact:
        return exact
    catalog = list(dict.fromkeys(by_compact.values()))
    if not clean_text(value) or not catalog:
        return None
    ranked = sorted(
        ((activity_affinity(value, name), name) for name in catalog),
        key=lambda item: (-item[0], key_text(item[1])),
    )
    best_score, best_name = ranked[0]
    second_score = ranked[1][0] if len(ranked) > 1 else 0.0
    return best_name if best_score >= 0.88 and best_score - second_score >= 0.12 else None


def activity_tokens(value: Any) -> set[str]:
    """Palabras significativas para comparar encabezados humanos del Excel."""
    return set(re.findall(r"[a-z0-9]+", key_text(value)))


def activity_affinity(candidate: Any, activity: Any) -> float:
    """Mide cercanía por escritura y palabras, sin depender de mayúsculas."""
    candidate_key = compact_key(candidate)
    activity_key = compact_key(activity)
    if not candidate_key or not activity_key:
        return 0.0
    character_score = SequenceMatcher(None, candidate_key, activity_key).ratio()
    candidate_words = activity_tokens(candidate)
    activity_words = activity_tokens(activity)
    if not candidate_words or not activity_words:
        return character_score
    overlap = len(candidate_words & activity_words)
    jaccard = overlap / len(candidate_words | activity_words)
    containment = overlap / min(len(candidate_words), len(activity_words))
    word_score = (0.65 * containment + 0.35 * jaccard) if min(
        len(candidate_words), len(activity_words)
    ) >= 2 else 0.0
    return max(character_score, word_score)


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for block in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def source_fingerprints(paths: dict[str, Path]) -> dict[str, str]:
    """Captura una versión estable de cada fuente antes de procesarla."""
    return {name: file_sha256(path) for name, path in paths.items()}


def ensure_source_stability(before: dict[str, str], paths: dict[str, Path]) -> None:
    """Evita publicar una mezcla si un Excel cambia durante la ejecución."""
    after = source_fingerprints(paths)
    changed = sorted(name for name, digest in before.items() if after.get(name) != digest)
    if changed:
        raise RuntimeError("Las fuentes cambiaron durante la actualización: " + ", ".join(changed))


def normalize_allowed_hosts(value: Any) -> set[str]:
    """Valida la lista CMS de dominios autorizados sin aceptar rutas ni esquemas."""
    hosts = set()
    for candidate in setting_list(value):
        host = candidate.casefold().rstrip(".")
        if not re.fullmatch(r"[a-z0-9](?:[a-z0-9.-]{0,251}[a-z0-9])?", host) or ".." in host:
            raise ValueError(f"Dominio de evidencia inválido en CMS: {candidate}")
        hosts.add(host)
    if not hosts:
        raise ValueError("El CMS debe definir al menos un dominio de evidencia autorizado")
    return hosts


def validate_xlsx(path: Path, label: str) -> None:
    """Rechaza libros incompletos/corruptos antes de que openpyxl los procese."""
    if not path.is_file():
        raise ValueError(f"No existe {label}: {path}")
    if path.suffix.casefold() != ".xlsx":
        raise ValueError(f"{label} debe ser un archivo .xlsx: {path.name}")
    try:
        with zipfile.ZipFile(path) as archive:
            bad_member = archive.testzip()
            missing = REQUIRED_XLSX_MEMBERS.difference(archive.namelist())
    except zipfile.BadZipFile as exc:
        raise ValueError(f"{label} no es un XLSX válido o está dañado: {path.name}") from exc
    if bad_member:
        raise ValueError(f"{label} contiene un componente dañado: {bad_member}")
    if missing:
        raise ValueError(f"{label} está incompleto: faltan {', '.join(sorted(missing))}")


def is_yes(value: Any) -> bool:
    return key_text(value) in {"si", "true", "1", "yes"}


def is_no(value: Any) -> bool:
    return key_text(value) in {"no", "false", "0"}


def boolean_answer(value: Any) -> bool | None:
    """Interpreta respuestas Sí/No, incluso cuando Forms agrega una aclaración."""
    normalized = key_text(value)
    if normalized in {"true", "1"} or re.match(r"^(?:si|yes)\b", normalized):
        return True
    if normalized in {"false", "0"} or re.match(r"^no\b", normalized):
        return False
    return None


def setting_list(value: Any) -> list[str]:
    """Normaliza una lista JSON o una lista separada por comas desde el CMS."""
    if isinstance(value, list):
        return [clean_text(item) for item in value if clean_text(item)]
    return [clean_text(item) for item in str(value or "").split(",") if clean_text(item)]


def normalize_ceco(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        value = str(int(value))
    text = clean_text(value).replace(".0", "")
    digits = re.sub(r"\D", "", text)
    return digits if len(digits) == 5 else ""


def evidence_key(activity: str, ceco: str) -> str:
    """Crea una etiqueta estable para identificar la actividad y el CeCo."""
    normalized = unicodedata.normalize("NFD", clean_text(activity))
    ascii_text = "".join(char for char in normalized if unicodedata.category(char) != "Mn")
    activity_token = re.sub(r"[^A-Za-z0-9]+", "_", ascii_text).strip("_") or "Evidencia"
    return f"{activity_token}_{ceco or 'CeCo_invalido'}"


def stable_response_id(*values: Any) -> str:
    """Identificador técnico estable; no depende de la columna Id de Forms."""
    raw = "|".join(clean_text(value) for value in values)
    return "respuesta-" + hashlib.sha256(raw.encode("utf-8")).hexdigest()[:16]


def safe_evidence_url(value: Any, allowed_hosts: set[str]) -> str | None:
    """Valida HTTPS y dominio, conservando el vínculo exactamente como llegó."""
    raw = str(value or "").strip()
    if not raw or len(raw) > 2048 or any(char in raw for char in ("\r", "\n", "\t")):
        return None
    try:
        parsed = urlsplit(raw)
        host = (parsed.hostname or "").casefold().rstrip(".")
        if (
            parsed.scheme.casefold() != "https"
            or not host
            or host not in allowed_hosts
            or parsed.username
            or parsed.password
            or parsed.port not in (None, 443)
        ):
            return None
        return raw
    except ValueError:
        return None


def evidence_filename(url: str | None) -> str:
    """Obtiene el nombre real del archivo desde el último segmento de la URL."""
    if not url:
        return "Sin archivo"
    name = unquote(urlsplit(url).path.rsplit("/", 1)[-1]).strip()
    return re.sub(r"[\r\n\t]", "", name) or "Sin archivo"


def parse_datetime(value: Any) -> datetime | None:
    if isinstance(value, datetime):
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time())
    text = clean_text(value)
    for fmt in ("%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def resolve_columns(headers: list[Any], contract: dict[str, tuple[str, ...]]) -> dict[str, int]:
    normalized = {key_text(value): index for index, value in enumerate(headers) if clean_text(value)}
    result: dict[str, int] = {}
    missing = []
    for field, aliases in contract.items():
        match = next((normalized[key_text(alias)] for alias in aliases if key_text(alias) in normalized), None)
        if match is None:
            missing.append(aliases[0])
        else:
            result[field] = match
    if missing:
        raise ValueError("Faltan encabezados requeridos: " + ", ".join(missing))
    return result


def matching_columns(headers: list[Any], aliases: tuple[str, ...]) -> list[int]:
    """Devuelve todas las columnas equivalentes, incluso si Forms las duplicó."""
    alias_keys = {key_text(alias) for alias in aliases}
    result = []
    for index, header in enumerate(headers):
        current = key_text(header)
        if not current:
            continue
        # Excel puede añadir sufijos .1 o (2) al desambiguar encabezados repetidos.
        without_duplicate_suffix = re.sub(r"(?:\s*[.(]\s*\d+\s*\)?|\.\d+)$", "", current).strip()
        if current in alias_keys or without_duplicate_suffix in alias_keys:
            result.append(index)
    return result


def closest_activity_key(candidate: Any, activity_names: list[str] | None) -> tuple[str | None, str | None]:
    """Relaciona un encabezado con el nombre CMS sin depender de posición.

    Acepta coincidencia exacta o una única coincidencia aproximada clara. Si dos
    actividades se parecen demasiado, falla de forma segura en vez de mezclar evidencia.
    """
    compact = compact_key(candidate)
    catalog = list(dict.fromkeys(
        (compact_key(name), clean_text(name))
        for name in (activity_names or []) if compact_key(name)
    ))
    if not compact or not catalog:
        return None, None
    keys = [key for key, _ in catalog]
    if compact in keys:
        return compact, "exact"
    ranked = sorted(
        ((activity_affinity(candidate, name), key) for key, name in catalog),
        reverse=True,
    )
    best_score, best_key = ranked[0]
    second_score = ranked[1][0] if len(ranked) > 1 else 0.0
    if best_score >= 0.78 and best_score - second_score >= 0.08:
        return best_key, "affinity"
    if best_score >= 0.78:
        return None, "ambiguous"
    return None, None


def evidence_header_activity(header: Any, activity_names: list[str] | None = None) -> tuple[str | None, str | None]:
    """Obtiene la actividad de Evidencia_<Actividad> o del nombre de actividad."""
    raw = key_text(header)
    if raw.startswith("evidencia"):
        if raw in {key_text(item) for item in EVIDENCE_HEADERS} or raw.startswith("evidencia del avance (pregunta"):
            return None, "generic"
        suffix = re.sub(r"^evidencia(?:\s+del\s+avance)?", "", raw, count=1)
        suffix = re.sub(r"\b(?:pregunta\s+no\s+anonima|respuesta\s+necesaria|cargar\s+archivo)\b", "", suffix)
        matched, match_type = closest_activity_key(suffix, activity_names)
        return (matched or compact_key(suffix) or None), (match_type or "unverified")
    matched, match_type = closest_activity_key(raw, activity_names)
    return matched, match_type


def evidence_columns(headers: list[Any], activity_names: list[str] | None = None) -> list[dict[str, Any]]:
    result = []
    for index, header in enumerate(headers):
        activity_key, match_type = evidence_header_activity(header, activity_names)
        raw_header = clean_text(header)
        is_question = "?" in raw_header or "¿" in raw_header
        # Una pregunta Sí/No puede contener exactamente el nombre de la actividad,
        # pero nunca debe confundirse con una columna para cargar archivos.
        if not is_question and (
            key_text(header).startswith("evidencia") or activity_key or match_type == "ambiguous"
        ):
            result.append({
                "index": index,
                "header": clean_text(header),
                "activityKey": activity_key,
                "matchType": match_type,
            })
    return result


def applicability_columns(
    headers: list[Any],
    excluded_indices: set[int],
) -> list[dict[str, Any]]:
    """Localiza preguntas operativas Sí/No sin depender de su posición."""
    result = []
    for index, header in enumerate(headers):
        if index in excluded_indices:
            continue
        raw = clean_text(header)
        normalized = key_text(header)
        if not raw:
            continue
        is_question = "?" in raw or "¿" in raw
        has_hint = any(hint in normalized for hint in APPLICABILITY_HEADER_HINTS)
        if is_question or has_hint:
            result.append({"index": index, "header": raw})
    return result


def resolve_applicability_answer(
    row: tuple[Any, ...],
    columns: list[dict[str, Any]],
) -> tuple[bool | None, list[str], str | None]:
    """Consolida Sí/No duplicados; otras preguntas se ignoran de forma segura."""
    answers: list[tuple[bool, str]] = []
    for column in columns:
        index = column["index"]
        value = clean_text(row[index]) if index < len(row) else ""
        if not value:
            continue
        answer = boolean_answer(value)
        if answer is None:
            continue
        answers.append((answer, column["header"]))
    if not answers:
        return None, [], None
    unique = {answer for answer, _ in answers}
    sources = list(dict.fromkeys(header for _, header in answers))
    if len(unique) > 1:
        return None, sources, "conflicting-applicability-answers"
    return answers[0][0], sources, None


def coalesce_row_value(row: tuple[Any, ...], indices: list[int]) -> tuple[str, bool]:
    """Une columnas equivalentes. Si hay valores distintos, no adivina."""
    values = []
    for index in indices:
        value = clean_text(row[index]) if index < len(row) else ""
        if value and value not in values:
            values.append(value)
    return (values[0] if len(values) == 1 else "", len(values) > 1)


def resolve_evidence_value(
    row: tuple[Any, ...],
    columns: list[dict[str, Any]],
    activity: str,
) -> tuple[str, str, str | None]:
    """Selecciona la evidencia por actividad y reporta ambigüedades sin mezclar archivos."""
    populated = []
    for column in columns:
        index = column["index"]
        value = clean_text(row[index]) if index < len(row) else ""
        if value:
            populated.append({**column, "value": value})
    if not populated:
        return "", "", None

    if any(item.get("matchType") == "ambiguous" for item in populated):
        return "", "", "ambiguous-evidence-header"

    activity_key = compact_key(activity)
    exact = [item for item in populated if item["activityKey"] == activity_key]
    exact_values = list(dict.fromkeys(item["value"] for item in exact))
    all_values = list(dict.fromkeys(item["value"] for item in populated))
    if len(exact_values) == 1:
        issue = "multiple-evidence-columns" if len(all_values) > 1 else None
        source = next(item["header"] for item in exact if item["value"] == exact_values[0])
        return exact_values[0], source, issue
    if len(exact_values) > 1:
        return "", "", "ambiguous-matching-evidence"

    generic = [item for item in populated if item["activityKey"] is None]
    generic_values = list(dict.fromkeys(item["value"] for item in generic))
    if len(generic_values) == 1 and len(all_values) == 1:
        source = next(item["header"] for item in generic if item["value"] == generic_values[0])
        return generic_values[0], source, "generic-evidence-fallback"
    if len(all_values) == 1:
        # Una evidencia en la columna de otra actividad no se reasigna por inferencia.
        return "", "", "mismatched-evidence-column"
    return "", "", "ambiguous-evidence"


def load_settings(path: Path, cms_settings: dict[str, Any] | None = None) -> dict[str, Any]:
    settings = json.loads(path.read_text(encoding="utf-8"))
    settings.update(cms_settings or {})
    return settings


def parse_date(value: Any):
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Fecha CMS inválida: {text}")


def date_status(start, end) -> str:
    today = datetime.now().date()
    if start and today < start:
        return "Programada"
    if end and today > end:
        return "Vencida"
    return "Vigente"


def find_header(ws, required: set[str]) -> tuple[int, dict[str, int]]:
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=12, values_only=True), 1):
        positions: dict[str, list[int]] = defaultdict(list)
        for index, value in enumerate(row):
            if clean_text(value):
                positions[key_text(value)].append(index)
        normalized = {key: indices[0] for key, indices in positions.items()}
        if required.issubset(normalized):
            duplicated = sorted(key for key in required if len(positions[key]) > 1)
            if duplicated:
                raise ValueError(
                    f"Encabezados CMS duplicados en {ws.title}: " + ", ".join(duplicated)
                )
            return row_number, normalized
    raise ValueError(f"No se encontró encabezado {', '.join(sorted(required))} en {ws.title}")


def load_cms(path: Path) -> tuple[list[dict[str, Any]], dict[str, dict[str, str]], dict[str, Any], dict[str, int]]:
    """Lee actividades, fechas, gerentes y configuración desde un solo Excel CMS."""
    validate_xlsx(path, "el CMS maestro")
    workbook = load_workbook(path, read_only=True, data_only=False)
    required_sheets = {"Actividades", "Gerentes", "Configuracion", "Organigrama"}
    missing = required_sheets.difference(workbook.sheetnames)
    if missing:
        raise ValueError("Faltan hojas CMS: " + ", ".join(sorted(missing)))

    config_ws = workbook["Configuracion"]
    config_header, config_cols = find_header(config_ws, {"clave", "valor"})
    cms_settings: dict[str, Any] = {}
    boolean_keys = {"onlyOpenStores", "requireEvidence", "publishEvidenceLinks", "publishPersonalData"}
    canonical_setting_keys = {key_text(key): key for key in KNOWN_SETTING_KEYS}
    seen_setting_keys: set[str] = set()
    for row in config_ws.iter_rows(min_row=config_header + 1, values_only=True):
        raw_key = clean_text(row[config_cols["clave"]])
        if not raw_key:
            continue
        normalized_key = key_text(raw_key)
        if normalized_key in seen_setting_keys:
            raise ValueError(f"El CMS contiene una clave de configuración duplicada: {raw_key}")
        seen_setting_keys.add(normalized_key)
        key = canonical_setting_keys.get(normalized_key, raw_key)
        value = row[config_cols["valor"]]
        text_value = clean_text(value)
        if not text_value:
            # Una celda borrada conserva el respaldo de config/settings.json.
            continue
        if key in boolean_keys:
            if not (is_yes(value) or is_no(value)):
                # Una edición incompleta no desactiva funciones por accidente.
                continue
            cms_settings[key] = is_yes(value)
        else:
            cms_settings[key] = text_value

    activity_ws = workbook["Actividades"]
    header_row, cols = find_header(activity_ws, {"orden", "actividad", "descripcion", "fecha inicio", "fecha limite", "activo"})
    activities = []
    calendar = {"active": 0, "scheduled": 0, "expired": 0, "inactive": 0}
    for row_number, row in enumerate(
        activity_ws.iter_rows(min_row=header_row + 1, values_only=True),
        header_row + 1,
    ):
        name = clean_text(row[cols["actividad"]])
        if not name:
            continue
        active_value = row[cols["activo"]]
        active = is_yes(active_value)
        if not active:
            # Filas nuevas, borradores o actividades marcadas No no bloquean el CMS.
            calendar["inactive"] += 1
            continue
        try:
            start = parse_date(row[cols["fecha inicio"]])
        except ValueError:
            start = None
        try:
            end = parse_date(row[cols["fecha limite"]])
        except ValueError:
            end = None
        if start and end and end < start:
            # Mantiene visible la actividad y evita publicar una fecha contradictoria.
            end = None
        status = date_status(start, end)
        if status == "Programada":
            calendar["scheduled"] += 1
        elif status == "Vencida":
            calendar["expired"] += 1
        calendar["active"] += 1
        evidence_col = cols.get("evidencia requerida")
        priority_col = cols.get("prioridad")
        evidence_value = row[evidence_col] if evidence_col is not None else None
        priority = clean_text(row[priority_col]) if priority_col is not None else ""
        try:
            order = int(float(row[cols["orden"]]))
        except (TypeError, ValueError):
            order = row_number
        activities.append({
            "name": name,
            "description": clean_text(row[cols["descripcion"]]),
            "order": order,
            "startDate": start.isoformat() if start else None,
            "endDate": end.isoformat() if end else None,
            "commitmentDateDisplay": end.strftime("%d/%m/%y") if end else "Sin fecha compromiso",
            "dateStatus": status,
            "requireEvidence": (
                is_yes(evidence_value)
                if is_yes(evidence_value) or is_no(evidence_value)
                else True
            ),
            "priority": priority or "Media",
            "autoDetected": False,
        })

    if not activities:
        raise ValueError("El CMS no contiene actividades activas para publicar")
    active_activity_catalog(activities)

    manager_ws = workbook["Gerentes"]
    manager_header, manager_cols = find_header(manager_ws, {"dm", "nombre corto", "foto webp", "activo"})
    managers: dict[str, dict[str, str]] = {}
    for row in manager_ws.iter_rows(min_row=manager_header + 1, values_only=True):
        dm = clean_text(row[manager_cols["dm"]])
        if not dm or not is_yes(row[manager_cols["activo"]]):
            continue
        manager_key = key_text(dm)
        if manager_key in managers:
            raise ValueError(f"El CMS contiene un gerente activo duplicado: {dm}")
        photo = clean_text(row[manager_cols["foto webp"]])
        if photo and not (ROOT / photo).is_file():
            raise ValueError(f"No existe la fotografía configurada para {dm}: {photo}")
        managers[manager_key] = {
            "shortName": clean_text(row[manager_cols["nombre corto"]]) or dm,
            "photo": photo,
        }
    org_ws = workbook["Organigrama"]
    org_header, org_cols = find_header(org_ws, {"nivel", "region", "nombre", "rol", "foto webp", "activo", "orden"})
    organization_rows = []
    for row_number, row in enumerate(org_ws.iter_rows(min_row=org_header + 1, values_only=True), org_header + 1):
        name = clean_text(row[org_cols["nombre"]])
        if not name or not is_yes(row[org_cols["activo"]]):
            continue
        photo = clean_text(row[org_cols["foto webp"]])
        if photo and not (ROOT / photo).is_file():
            raise ValueError(f"No existe la fotografía del organigrama para {name}: {photo}")
        try:
            level = int(float(row[org_cols["nivel"]]))
        except (TypeError, ValueError):
            level = 2
        try:
            order = int(float(row[org_cols["orden"]]))
        except (TypeError, ValueError):
            order = row_number
        organization_rows.append({
            "level": level,
            "region": clean_text(row[org_cols["region"]]),
            "name": name,
            "role": clean_text(row[org_cols["rol"]]),
            "photo": photo,
            "order": order,
        })
    organization_rows.sort(key=lambda item: (item["level"], item["order"], key_text(item["name"])))
    national = next((item for item in organization_rows if item["level"] == 1), None)
    if not national:
        raise ValueError("El CMS debe incluir un Director Starbucks México activo en Organigrama")
    organization = {
        "scopeLabel": "Región | Centro's",
        "nationalDirector": national,
        "regionalDirectors": [item for item in organization_rows if item["level"] == 2],
    }
    cms_settings["_organization"] = organization
    return sorted(activities, key=lambda item: (item["order"], key_text(item["name"]))), managers, cms_settings, calendar


def status_label(compliance: float) -> str:
    if compliance >= 80:
        return "En meta"
    if compliance >= 40:
        return "Seguimiento"
    return "Atención"


def deadline_focus(end_date: str | None, pending: int) -> dict[str, Any]:
    """Define el foco visual desde la fecha CMS sin ocultar actividades activas."""
    if pending <= 0:
        return {"deadlineLabel": "Completa", "deadlineTone": "green", "daysRemaining": None}
    if not end_date:
        return {"deadlineLabel": "Sin fecha", "deadlineTone": "neutral", "daysRemaining": None}
    end = datetime.fromisoformat(end_date).date()
    days = (end - datetime.now().date()).days
    if days < 0:
        return {"deadlineLabel": "Vencida", "deadlineTone": "red", "daysRemaining": days}
    if days == 0:
        return {"deadlineLabel": "Vence hoy", "deadlineTone": "red", "daysRemaining": 0}
    if days <= 7:
        return {"deadlineLabel": f"Vence en {days} días", "deadlineTone": "amber", "daysRemaining": days}
    return {"deadlineLabel": f"En {days} días", "deadlineTone": "green", "daysRemaining": days}


def find_directory_header(ws) -> tuple[int, list[Any]]:
    for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=min(ws.max_row, 12), values_only=True), 1):
        keys = {key_text(value) for value in row if clean_text(value)}
        if {"cc", "cc nombre", "dm"}.issubset(keys):
            return row_number, list(row)
    raise ValueError(f"No se encontró encabezado CC / CC Nombre / DM en {ws.title}")


def is_all_regions(value: Any) -> bool:
    return key_text(value) in {"", "todas", "todos", "todas las regiones", "*"}


def included_store_statuses(settings: dict[str, Any]) -> set[str]:
    """Devuelve el alcance CMS; si no se indica, el único estatus válido es Abierta."""
    configured = setting_list(settings.get("includedStoreStatuses", "Abierta"))
    statuses = {key_text(value) for value in configured if clean_text(value)}
    return statuses or {"abierta"}


def directory_sheet(workbook, requested: Any) -> tuple[Any, int, list[Any]]:
    """Elige por encabezados, no por un nombre o un número de filas congelado."""
    candidates = []
    requested_name = clean_text(requested)
    for index, ws in enumerate(workbook.worksheets):
        try:
            header_row, headers = find_directory_header(ws)
        except ValueError:
            continue
        keys = {key_text(value) for value in headers if clean_text(value)}
        score = sum(field in keys for field in ("cc", "cc nombre", "region", "dm", "estatus"))
        preferred = int(ws.title == requested_name)
        candidates.append((preferred, score, ws.max_row, -index, ws, header_row, headers))
    if not candidates:
        raise ValueError("No existe una hoja con encabezados CC / CC Nombre / Región / DM")
    _, _, _, _, ws, header_row, headers = max(candidates, key=lambda item: item[:4])
    return ws, header_row, headers


def load_directory(path: Path, settings: dict[str, Any]) -> tuple[dict[str, dict[str, str]], str, dict[str, Any]]:
    validate_xlsx(path, "el directorio")
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws, header_row, headers = directory_sheet(workbook, settings.get("directorySheet"))
    required = ("cc", "cc nombre", "region", "dm")
    positions: dict[str, list[int]] = defaultdict(list)
    for index, value in enumerate(headers):
        if clean_text(value):
            positions[key_text(value)].append(index)
    duplicated_headers = sorted(field for field in required if len(positions.get(field, [])) > 1)
    if duplicated_headers:
        raise ValueError("Directorio con encabezados duplicados: " + ", ".join(duplicated_headers))
    normalized = {key: indices[0] for key, indices in positions.items()}
    missing = [field for field in required if field not in normalized]
    if missing:
        raise ValueError("Directorio incompleto: " + ", ".join(missing))

    status_index = normalized.get("estatus")
    filter_by_status = bool(settings.get("onlyOpenStores"))
    allowed_statuses = included_store_statuses(settings)
    if filter_by_status and status_index is None:
        raise ValueError("El Directorio debe incluir la columna Estatus para publicar sólo tiendas abiertas")

    stores: dict[str, dict[str, str]] = {}
    status_counts: Counter[str] = Counter()
    excluded_status_counts: Counter[str] = Counter()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        ceco = normalize_ceco(row[normalized["cc"]])
        if not ceco:
            continue
        region = clean_text(row[normalized["region"]])
        status = clean_text(row[status_index]) if status_index is not None else "Sin estatus en fuente"
        status_counts[status] += 1
        if not is_all_regions(settings.get("region")) and key_text(region) != key_text(settings["region"]):
            continue
        if filter_by_status and key_text(status) not in allowed_statuses:
            excluded_status_counts[status] += 1
            continue
        if ceco in stores:
            raise ValueError(f"CeCo duplicado en el directorio operativo: {ceco}")
        stores[ceco] = {
            "ceco": ceco,
            "store": clean_text(row[normalized["cc nombre"]]) or f"Tienda {ceco}",
            "dm": normalize_dm(row[normalized["dm"]]),
            "region": region,
            "status": status,
        }
    if not stores:
        raise ValueError("El directorio no contiene tiendas activas para el alcance configurado")
    return stores, ws.title, {
        "includedStatuses": sorted(
            {status for status in status_counts if key_text(status) in allowed_statuses}, key=key_text
        ),
        "includedStores": len(stores),
        "excludedStores": sum(excluded_status_counts.values()),
        "sourceStatusCounts": dict(sorted(status_counts.items(), key=lambda item: key_text(item[0]))),
        "excludedStatusCounts": dict(sorted(excluded_status_counts.items(), key=lambda item: key_text(item[0]))),
    }


def find_response_source(workbook, activity_names: list[str] | None = None) -> tuple[Any, int, list[Any]]:
    """Localiza hoja y fila de encabezados aunque Forms agregue portada o filas previas."""
    candidates = []
    for sheet_index, ws in enumerate(workbook.worksheets):
        scan_limit = min(max(ws.max_row, 1), 25)
        for row_number, row in enumerate(ws.iter_rows(min_row=1, max_row=scan_limit, values_only=True), 1):
            headers = list(row)
            activity_columns = matching_columns(headers, RESPONSE_FIELDS["activity"])
            ceco_columns = matching_columns(headers, RESPONSE_FIELDS["ceco"])
            evidence_group = evidence_columns(headers, activity_names)
            if activity_columns and ceco_columns and evidence_group:
                score = len(evidence_group) * 100 + sum(
                    bool(matching_columns(headers, aliases)) for aliases in RESPONSE_FIELDS.values()
                )
                candidates.append((score, -sheet_index, -row_number, ws, row_number, headers))
    if not candidates:
        raise ValueError("No se encontró una tabla Forms con Actividad, CeCo y Evidencia")
    _, _, _, ws, header_row, headers = max(candidates, key=lambda item: item[:3])
    return ws, header_row, headers


def load_responses(path: Path, activity_names: list[str] | None = None) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    """Lee exportaciones Forms antiguas, anchas o normalizadas por filas.

    Soporta una sola columna genérica de evidencia, múltiples columnas
    Evidencia_<Actividad>, encabezados duplicados y columnas reordenadas.
    """
    validate_xlsx(path, "la exportación de Forms")
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws, header_row, headers = find_response_source(workbook, activity_names)
    rows = ws.iter_rows(min_row=header_row + 1, values_only=True)
    column_groups = {field: matching_columns(headers, aliases) for field, aliases in RESPONSE_FIELDS.items()}
    missing = [RESPONSE_FIELDS[field][0] for field in REQUIRED_RESPONSE_FIELDS if not column_groups[field]]
    if missing:
        raise ValueError("Faltan encabezados requeridos: " + ", ".join(missing))
    confirmation_columns = matching_columns(headers, CONFIRMATION_HEADERS)
    evidence_group = evidence_columns(headers, activity_names)
    if not evidence_group:
        raise ValueError("No se encontró ninguna columna de evidencia")
    excluded_indices = {
        index for indices in column_groups.values() for index in indices
    } | set(confirmation_columns) | {item["index"] for item in evidence_group}
    applicability_group = applicability_columns(headers, excluded_indices)
    response_activity_by_text: dict[str, str] = {}
    response_activity_by_compact: dict[str, str] = {}
    if activity_names:
        response_activity_by_text, response_activity_by_compact = active_activity_catalog(
            [{"name": name} for name in activity_names]
        )

    responses = []
    conflicts = []
    evidence_issues: dict[str, list[int]] = defaultdict(list)
    applicability_issues: dict[str, list[int]] = defaultdict(list)
    for row_number, row in enumerate(rows, header_row + 1):
        if not any(value not in (None, "") for value in row):
            continue
        values: dict[str, str] = {}
        row_has_conflict = False
        for field, indices in column_groups.items():
            values[field], conflict = coalesce_row_value(row, indices)
            if conflict:
                conflicts.append({"row": row_number, "field": field})
                row_has_conflict = True
        confirmation, confirmation_conflict = coalesce_row_value(row, confirmation_columns)
        if confirmation_conflict:
            conflicts.append({"row": row_number, "field": "confirmed"})
            row_has_conflict = True
        evidence_activity = canonical_cms_activity(
            values["activity"], response_activity_by_text, response_activity_by_compact
        ) or values["activity"]
        evidence, evidence_source, evidence_issue = resolve_evidence_value(
            row, evidence_group, evidence_activity
        )
        if evidence_issue:
            evidence_issues[evidence_issue].append(row_number)
        applicability, applicability_sources, applicability_issue = resolve_applicability_answer(
            row, applicability_group
        )
        if applicability_issue:
            applicability_issues[applicability_issue].append(row_number)
            row_has_conflict = True
        finished = parse_datetime(values["finished"])
        # Registrar una actividad en Forms equivale a confirmarla. La respuesta de
        # confirmación puede permanecer en exportaciones históricas, pero nunca
        # cambia aplicabilidad ni cumplimiento. La evidencia sigue siendo obligatoria
        # cuando así lo define el CMS.
        confirmed = bool(values["activity"])
        response_id = stable_response_id(
            values["started"], values["finished"], values["ceco"], values["activity"], evidence
        )
        responses.append({
            "row": row_number,
            "id": response_id,
            "started": parse_datetime(values["started"]),
            "finished": finished,
            "email": values["email"],
            "name": values["name"],
            "activity": values["activity"],
            "ceco": normalize_ceco(values["ceco"]),
            "confirmedAnswer": "Sí" if values["activity"] else "",
            "confirmed": confirmed and not row_has_conflict,
            "applicabilityAnswer": "Sí" if applicability is True else ("No" if applicability is False else ""),
            "applicabilitySourceHeaders": applicability_sources,
            "applicabilityConflict": applicability_issue is not None,
            "explicitNo": applicability is False and not applicability_issue,
            "evidence": evidence,
            "evidenceSourceHeader": evidence_source,
            "schemaConflict": row_has_conflict,
            "evidenceIssue": evidence_issue,
        })
    schema = {
        "sheet": ws.title,
        "headerRow": header_row,
        "columns": len(headers),
        "activityHeaders": [clean_text(headers[index]) for index in column_groups["activity"]],
        "cecoHeaders": [clean_text(headers[index]) for index in column_groups["ceco"]],
        "confirmationHeaders": [clean_text(headers[index]) for index in confirmation_columns],
        "applicabilityHeaders": [item["header"] for item in applicability_group],
        "evidenceHeaders": [item["header"] for item in evidence_group],
        "evidenceHeaderMap": {
            item["header"]: item["activityKey"] or "generic"
            for item in evidence_group
        },
        "evidenceHeaderMatch": {
            item["header"]: item.get("matchType") or "none"
            for item in evidence_group
        },
        "rowConflicts": conflicts,
        "evidenceIssues": dict(evidence_issues),
        "applicabilityIssues": dict(applicability_issues),
    }
    return responses, schema


def iso_or_none(value: datetime | None) -> str | None:
    return value.isoformat(timespec="seconds") if value else None


def build_payload(
    responses_path: Path,
    directory_path: Path,
    settings_path: Path,
    cms_path: Path = DEFAULT_CMS,
) -> dict[str, Any]:
    source_paths = {
        "responsesSha256": responses_path,
        "directorySha256": directory_path,
        "cmsSha256": cms_path,
    }
    initial_source_hashes = source_fingerprints(source_paths)
    activities, managers, cms_settings, calendar = load_cms(cms_path)
    organization = cms_settings.pop("_organization")
    settings = load_settings(settings_path, cms_settings)
    allowed_hosts = normalize_allowed_hosts(
        settings.get("evidenceAllowedHosts", "grupovips-my.sharepoint.com")
    )
    regional_director_photo = clean_text(settings.get("regionalDirectorPhoto", "assets/director/jorge-alcantar.webp"))
    if regional_director_photo and not (ROOT / regional_director_photo).is_file():
        raise ValueError(f"No existe la fotografía del Director Regional: {regional_director_photo}")
    stores, directory_sheet, directory_status = load_directory(directory_path, settings)
    responses, response_schema = load_responses(responses_path, [item["name"] for item in activities])

    # El Forms acumula historia; sólo el CMS decide qué actividades se publican.
    configured_by_text, configured_by_compact = active_activity_catalog(activities)
    activity_names = [item["name"] for item in activities]
    evidence_rules = {key_text(item["name"]): item.get("requireEvidence", True) for item in activities}

    submissions = []
    latest_state_by_pair: dict[tuple[str, str], dict[str, Any]] = {}
    conditional_activities: set[str] = set()
    unknown_cecos = set()
    invalid_rows = []
    unsafe_evidence_rows = []
    hidden_activity_rows = []
    hidden_activities = set()
    canonicalized_activity_rows = []
    latest_update = None

    for response in responses:
        store = stores.get(response["ceco"])
        activity_text = clean_text(response["activity"])
        activity = canonical_cms_activity(activity_text, configured_by_text, configured_by_compact)
        if not activity_text:
            invalid_rows.append(response["row"])
            continue
        if not activity:
            hidden_activity_rows.append(response["row"])
            hidden_activities.add(activity_text)
            continue
        if compact_key(activity_text) != compact_key(activity):
            canonicalized_activity_rows.append(response["row"])
        if response["ceco"] and not store:
            unknown_cecos.add(response["ceco"])
        # Una fila ajena o inactiva no modifica ni los conteos ni la fecha de corte.
        if response["finished"] and (latest_update is None or response["finished"] > latest_update):
            latest_update = response["finished"]
        evidence_url = safe_evidence_url(response["evidence"], allowed_hosts)
        evidence_available = evidence_url is not None
        not_applicable = bool(response["explicitNo"] and store and not response["applicabilityConflict"])
        if response["applicabilityAnswer"]:
            conditional_activities.add(activity)
        if response["evidence"] and not evidence_available:
            unsafe_evidence_rows.append(response["row"])
        valid = bool(
            store
            and activity
            and response["confirmed"]
            and not not_applicable
            and (evidence_available or not evidence_rules.get(key_text(activity), settings.get("requireEvidence", True)))
        )
        if not valid and not not_applicable:
            invalid_rows.append(response["row"])

        key = evidence_key(activity or "Evidencia", response["ceco"])
        public = {
            "id": response["id"],
            "timestamp": iso_or_none(response["finished"]),
            "timestampDisplay": response["finished"].strftime("%d/%m/%Y %H:%M") if response["finished"] else "Sin fecha",
            "activity": activity or "Sin actividad",
            "ceco": response["ceco"] or "Inválido",
            "store": store["store"] if store else "CeCo sin cruce",
            "dm": store["dm"] if store else "Sin asignar",
            "region": store["region"] if store else "Sin región",
            "evidenceKey": key,
            "evidenceLinkLabel": f"Link_{key}",
            "evidenceFileName": evidence_filename(evidence_url),
            "evidenceSourceHeader": response["evidenceSourceHeader"],
            "confirmed": response["confirmed"],
            "answer": response["confirmedAnswer"],
            "notApplicable": not_applicable,
            "status": "Realizada" if valid else "Pendiente",
            "evidenceAvailable": evidence_available,
            "evidenceLinkPublished": bool(settings.get("publishEvidenceLinks") and evidence_url),
            "valid": valid,
        }
        if settings.get("publishEvidenceLinks") and evidence_url:
            public["evidenceUrl"] = evidence_url
        if settings.get("publishPersonalData"):
            public["submittedBy"] = response["name"]
            public["email"] = response["email"]
        submissions.append(public)

        if store and activity and not response["applicabilityConflict"] and (
            valid or response["applicabilityAnswer"]
        ):
            pair = (response["ceco"], activity)
            state = {**response, "valid": valid, "notApplicable": not_applicable}
            current = latest_state_by_pair.get(pair)
            current_sort = ((current or {}).get("finished") or datetime.min, (current or {}).get("row", 0))
            state_sort = (response["finished"] or datetime.min, response["row"])
            if current is None or state_sort > current_sort:
                latest_state_by_pair[pair] = state

    not_applicable_pairs = {
        pair for pair, state in latest_state_by_pair.items() if state["notApplicable"]
    }
    latest_by_pair = {
        pair: state for pair, state in latest_state_by_pair.items()
        if state["valid"] and not state["notApplicable"]
    }
    completion_pairs = set(latest_by_pair)
    raw_valid_responses = sum(item["valid"] for item in submissions)
    latest_submission_by_pair: dict[tuple[str, str], dict[str, Any]] = {}
    for item in submissions:
        if not (item["valid"] or item["notApplicable"]):
            continue
        pair = (item["ceco"], item["activity"])
        current = latest_submission_by_pair.get(pair)
        item_sort = (item.get("timestamp") or "", item.get("id") or "")
        current_sort = ((current or {}).get("timestamp") or "", (current or {}).get("id") or "")
        if current is None or item_sort > current_sort:
            latest_submission_by_pair[pair] = item
    submissions = list(latest_submission_by_pair.values())
    latest_timestamp_by_ceco: dict[str, datetime] = {}
    for (ceco, _), item in latest_by_pair.items():
        finished = item.get("finished")
        if finished and (ceco not in latest_timestamp_by_ceco or finished > latest_timestamp_by_ceco[ceco]):
            latest_timestamp_by_ceco[ceco] = finished
    completed_by_activity = Counter(activity for _, activity in completion_pairs)
    not_applicable_by_activity = Counter(activity for _, activity in not_applicable_pairs)
    store_rows = []
    for ceco, store in sorted(stores.items(), key=lambda item: (key_text(item[1]["dm"]), key_text(item[1]["store"]))):
        status = {activity: (ceco, activity) in completion_pairs for activity in activity_names}
        applicability = {activity: (ceco, activity) not in not_applicable_pairs for activity in activity_names}
        completed = sum(status.values())
        expected = sum(applicability.values())
        not_applicable = len(activity_names) - expected
        store_rows.append({
            **store,
            "completed": completed,
            "expected": expected,
            "notApplicable": not_applicable,
            "compliance": round(completed / expected * 100, 1) if expected else 0,
            "lastUpdate": iso_or_none(latest_timestamp_by_ceco.get(ceco)),
            "activities": status,
            "applicableActivities": applicability,
        })

    activity_stats = []
    for item in activities:
        completed = completed_by_activity[item["name"]]
        not_applicable = not_applicable_by_activity[item["name"]]
        applicable = len(stores) - not_applicable
        pending = applicable - completed
        activity_stats.append({
            **item,
            "conditionalApplicability": item["name"] in conditional_activities,
            "completedStores": completed,
            "applicableStores": applicable,
            "notApplicableStores": not_applicable,
            "pendingStores": pending,
            "compliance": round(completed / applicable * 100, 1) if applicable else 0,
            **deadline_focus(item.get("endDate"), pending),
        })
    activity_stats.sort(key=lambda item: (
        item["pendingStores"] == 0,
        item.get("endDate") or "9999-12-31",
        item["order"],
        key_text(item["name"]),
    ))
    for focus_rank, item in enumerate(activity_stats, 1):
        item["focusRank"] = focus_rank

    dm_groups: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for store in store_rows:
        dm_groups[store["dm"]].append(store)
    dm_stats = []
    for dm, dm_stores in sorted(dm_groups.items(), key=lambda item: key_text(item[0])):
        completed = sum(store["completed"] for store in dm_stores)
        expected = sum(store["expected"] for store in dm_stores)
        not_applicable = sum(store["notApplicable"] for store in dm_stores)
        compliance = round(completed / expected * 100, 1) if expected else 0
        profile = managers.get(key_text(dm), {})
        pending_stores = sum(store["completed"] < store["expected"] for store in dm_stores)
        dm_stats.append({
            "dm": dm,
            "shortName": profile.get("shortName", dm),
            "photo": profile.get("photo", ""),
            "photoStatus": "Disponible" if profile.get("photo") else "Pendiente",
            "regions": sorted({store["region"] for store in dm_stores}, key=key_text),
            "stores": len(dm_stores),
            "completed": completed,
            "expected": expected,
            "notApplicable": not_applicable,
            "pending": expected - completed,
            "pendingStores": pending_stores,
            "compliance": compliance,
            "status": status_label(compliance),
        })
    dm_stats.sort(key=lambda item: (-item["compliance"], key_text(item["shortName"])))
    for rank, item in enumerate(dm_stats, 1):
        item["rank"] = rank

    not_applicable_total = len(not_applicable_pairs)
    expected_total = len(stores) * len(activity_names) - not_applicable_total
    completed_total = len(completion_pairs)
    valid_responses = sum(item["valid"] for item in submissions)
    stores_complete = sum(item["completed"] == item["expected"] and item["expected"] > 0 for item in store_rows)

    ensure_source_stability(initial_source_hashes, source_paths)
    source_hashes = initial_source_hashes
    version_inputs = dict(source_hashes)
    for relative_path in (
        "scripts/build_dashboard.py", "app.js", "styles.css", "service-worker.js",
        "pdf-export.js", "xlsx-export.js", "index.html",
    ):
        version_inputs[relative_path] = file_sha256(ROOT / relative_path)
    build_version = hashlib.sha256(
        json.dumps(version_inputs, sort_keys=True).encode("utf-8")
    ).hexdigest()[:16]

    ambiguous_evidence_issues = {
        "ambiguous-evidence", "ambiguous-matching-evidence",
        "ambiguous-evidence-header", "mismatched-evidence-column",
        "multiple-evidence-columns",
    }
    published_pairs = [(item["ceco"], item["activity"]) for item in submissions]
    stability_controls = {
        "sourceIntegrity": True,
        "cmsActiveAllowlist": bool(activity_names) and len(activity_names) == len(set(activity_names)),
        "externalFormsIsolation": all(item["activity"] in activity_names for item in submissions),
        "canonicalActivityMatching": all(item["activity"] in activity_names for item in submissions),
        "evidenceHeaderSafety": not any(
            rows for issue, rows in response_schema.get("evidenceIssues", {}).items()
            if issue in ambiguous_evidence_issues
        ),
        "columnOrderIndependence": bool(
            response_schema.get("activityHeaders")
            and response_schema.get("cecoHeaders")
            and response_schema.get("evidenceHeaders")
        ),
        "duplicateResponseResolution": len(published_pairs) == len(set(published_pairs)),
        "directoryUniqueness": bool(stores) and len(stores) == len(set(stores)),
        "safeEvidenceLinks": not unsafe_evidence_rows,
        "atomicPublication": True,
    }
    stability_passed = sum(stability_controls.values())

    regions = sorted({store["region"] for store in store_rows}, key=key_text)
    region_label = regions[0] if len(regions) == 1 else "Todas las regiones"

    return {
        "schemaVersion": 12,
        "buildVersion": build_version,
        "project": settings.get("projectName", "Sistema de Evidencias OPS"),
        "region": region_label,
        "regions": regions,
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "lastUpdated": iso_or_none(latest_update),
        "lastUpdatedDisplay": latest_update.strftime("%d/%m/%Y %H:%M") if latest_update else "Sin respuestas",
        "report": {
            "title": "Sistema de Evidencia OPS",
            "subtitle": "Dashboard de Avance de Actividades",
            "motto": "JUNTÉMONOS MÁS",
            "footerLabel": "Starbucks México · Operaciones",
            "cutOffDisplay": latest_update.strftime("%d/%m/%y · %H:%M h") if latest_update else "Sin datos",
            "regionalDirector": {
                "name": clean_text(settings.get("regionalDirectorName", "Jorge Alcantar")),
                "role": "Director Regional",
                "photo": regional_director_photo,
            },
        },
        "organization": organization,
        "sources": {
            "responses": responses_path.name,
            "responsesSha256": source_hashes["responsesSha256"],
            "directory": directory_path.name,
            "directorySha256": source_hashes["directorySha256"],
            "directorySheet": directory_sheet,
            "directoryStatus": directory_status,
            "cms": cms_path.name,
            "cmsSha256": source_hashes["cmsSha256"],
        },
        "summary": {
            "regions": len(regions),
            "dms": len(dm_stats),
            "stores": len(stores),
            "activities": len(activity_names),
            "expectedCompletions": expected_total,
            "completedCompletions": completed_total,
            "compliance": round(completed_total / expected_total * 100, 1) if expected_total else 0,
            "validResponses": valid_responses,
            "storesComplete": stores_complete,
            "pendingCompletions": expected_total - completed_total,
            "notApplicableCompletions": not_applicable_total,
        },
        "quality": {
            "responsesRead": len(responses),
            "invalidRows": invalid_rows,
            "unknownCeCos": sorted(unknown_cecos),
            "hiddenActivityRows": hidden_activity_rows,
            "hiddenActivities": sorted(hidden_activities, key=key_text),
            "canonicalizedActivityRows": canonicalized_activity_rows,
            "duplicateValidResponses": max(raw_valid_responses - valid_responses, 0),
            "unsafeEvidenceRows": unsafe_evidence_rows,
            "responseSchema": response_schema,
            "notApplicableResponses": sum(item["notApplicable"] for item in submissions),
            "notApplicablePairs": not_applicable_total,
            "evidenceLinksPublished": sum(bool(item.get("evidenceUrl")) for item in submissions),
            "privacyMode": not settings.get("publishPersonalData") and not settings.get("publishEvidenceLinks"),
            "stabilityControls": stability_controls,
            "stabilityScore": f"{stability_passed}/{len(STABILITY_CONTROLS)}",
        },
        "calendar": calendar,
        "activities": activity_stats,
        "dms": dm_stats,
        "attention": sorted(
            [
                {
                    "ceco": store["ceco"],
                    "store": store["store"],
                    "dm": store["dm"],
                    "completed": store["completed"],
                    "expected": store["expected"],
                    "notApplicable": store["notApplicable"],
                    "pending": store["expected"] - store["completed"],
                    "compliance": store["compliance"],
                    "status": status_label(store["compliance"]),
                }
                for store in store_rows
                if store["completed"] < store["expected"]
            ],
            key=lambda item: (item["compliance"], key_text(item["dm"]), key_text(item["store"])),
        ),
        "stores": store_rows,
        "submissions": sorted(submissions, key=lambda item: item["timestamp"] or "", reverse=True),
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera data/dashboard.json desde los Excel del proyecto.")
    parser.add_argument("--responses", type=Path, default=DEFAULT_RESPONSES)
    parser.add_argument("--directory", type=Path, default=DEFAULT_DIRECTORY)
    parser.add_argument("--settings", type=Path, default=DEFAULT_SETTINGS)
    parser.add_argument("--cms", type=Path, default=DEFAULT_CMS)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    payload = build_payload(args.responses, args.directory, args.settings, args.cms)
    atomic_write_text(
        args.output,
        json.dumps(payload, ensure_ascii=False, indent=2) + "\n",
    )
    summary = payload["summary"]
    print(
        f"Dashboard generado: {summary['stores']} tiendas · {summary['activities']} actividades · "
        f"{summary['completedCompletions']}/{summary['expectedCompletions']} cumplimientos"
    )
    print(f"Última actualización Forms: {payload['lastUpdatedDisplay']}")


if __name__ == "__main__":
    main()
