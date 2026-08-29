#!/usr/bin/env python3
"""Genera el resumen ejecutivo XLSX público del Sistema de Evidencias OPS."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DATA = ROOT / "data" / "dashboard.json"
DEFAULT_OUTPUT = ROOT / "exports" / "Resumen_Evidencias_OPS.xlsx"

GREEN = "006241"
DARK = "1E3932"
SOFT = "E9F4EF"
CANVAS = "F6F8F7"
LINE = "DCE5E0"
WHITE = "FFFFFF"
RED = "C54435"
AMBER = "C98612"
GOOD = "16845B"


def style_title(ws, end_column: int, title: str, subtitle: str) -> None:
    end = get_column_letter(end_column)
    ws.merge_cells(f"A1:{end}2")
    ws["A1"] = title
    ws["A1"].font = Font(name="Aptos Display", size=20, bold=True, color=WHITE)
    ws["A1"].fill = PatternFill("solid", fgColor=GREEN)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.merge_cells(f"A3:{end}3")
    ws["A3"] = subtitle
    ws["A3"].font = Font(name="Aptos", size=10, italic=True, color=DARK)
    ws["A3"].fill = PatternFill("solid", fgColor=SOFT)
    ws["A3"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 12
    ws.row_dimensions[3].height = 22
    ws.sheet_view.showGridLines = False


def style_header(ws, row: int, start_column: int, end_column: int) -> None:
    for cell in ws.iter_cols(min_col=start_column, max_col=end_column, min_row=row, max_row=row):
        item = cell[0]
        item.font = Font(name="Aptos", size=10, bold=True, color=WHITE)
        item.fill = PatternFill("solid", fgColor=DARK)
        item.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 24


def style_table(ws, start_row: int, end_row: int, end_column: int) -> None:
    bottom = Side(style="thin", color=LINE)
    for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=1, max_col=end_column):
        for cell in row:
            cell.font = Font(name="Aptos", size=10, color=DARK)
            cell.border = Border(bottom=bottom)
            cell.alignment = Alignment(vertical="center")
            if cell.row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=CANVAS)
        ws.row_dimensions[row[0].row].height = 21


def set_widths(ws, widths: list[float]) -> None:
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width


def status(value: float) -> str:
    if value >= 80:
        return "En meta"
    if value >= 40:
        return "Seguimiento"
    return "Atención"


def style_status_column(ws, start_row: int, end_row: int, column: int) -> None:
    palette = {
        "En meta": ("DAF1E6", "116444"),
        "Seguimiento": ("FFF0D5", "80520C"),
        "Atención": ("F9E6E3", "922F24"),
    }
    for row in range(start_row, end_row + 1):
        cell = ws.cell(row=row, column=column)
        fill, font = palette.get(cell.value, (CANVAS, DARK))
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(name="Aptos", size=9, bold=True, color=font)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def build_workbook(data: dict) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Resumen"
    stores_sheet = workbook.create_sheet("Tiendas")
    activities_sheet = workbook.create_sheet("Actividades")
    source = data.get("summary", {})
    cut = data.get("lastUpdatedDisplay", "Sin datos")
    region = data.get("region", "Centro Norte")

    style_title(summary, 8, "Sistema de Evidencias OPS", f"Resumen ejecutivo · Región {region} · Corte {cut}")
    summary.append([])
    summary.append(["Realizadas", None, "Total", None, "% Avance", None, "Pendientes", None])
    summary.append([source.get("completedCompletions", 0), None, source.get("expectedCompletions", 0), None, None, None, source.get("pendingCompletions", 0), None])
    summary.merge_cells("A5:B5"); summary.merge_cells("C5:D5"); summary.merge_cells("E5:F5"); summary.merge_cells("G5:H5")
    summary.merge_cells("A6:B7"); summary.merge_cells("C6:D7"); summary.merge_cells("E6:F7"); summary.merge_cells("G6:H7")
    for label_cell in ("A5", "C5", "E5", "G5"):
        summary[label_cell].fill = PatternFill("solid", fgColor=DARK)
        summary[label_cell].font = Font(name="Aptos", size=9, bold=True, color=WHITE)
        summary[label_cell].alignment = Alignment(horizontal="center", vertical="center")
    for value_cell in ("A6", "C6", "E6", "G6"):
        summary[value_cell].fill = PatternFill("solid", fgColor=SOFT)
        summary[value_cell].font = Font(name="Aptos Display", size=18, bold=True, color=GREEN)
        summary[value_cell].alignment = Alignment(horizontal="center", vertical="center")
    for value_cell in ("A6", "C6", "G6"):
        summary[value_cell].number_format = "#,##0"
    summary["E6"].number_format = "0.0%"

    headers = ["Ranking", "DM", "Tiendas", "Realizadas", "Total", "Pendientes", "% Avance", "Estado"]
    for column, value in enumerate(headers, 1):
        summary.cell(row=9, column=column, value=value)
    dms = sorted(data.get("dms", []), key=lambda item: (-item.get("compliance", 0), item.get("shortName", "")))
    for rank, item in enumerate(dms, 1):
        row = 9 + rank
        values = [rank, item.get("shortName"), item.get("stores", 0), item.get("completed", 0), item.get("expected", 0), item.get("pending", 0), f"=IFERROR(D{row}/E{row},0)", status(item.get("compliance", 0))]
        for column, value in enumerate(values, 1):
            summary.cell(row=row, column=column, value=value)
    summary["E6"] = f"=IFERROR(SUM(D10:D{summary.max_row})/SUM(E10:E{summary.max_row}),0)"
    style_header(summary, 9, 1, 8)
    style_table(summary, 10, summary.max_row, 8)
    summary.freeze_panes = "A10"
    summary.auto_filter.ref = f"A9:H{summary.max_row}"
    summary.conditional_formatting.add(f"G10:G{summary.max_row}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=GREEN, showValue=True))
    for row in range(10, summary.max_row + 1):
        for column in (1, 3, 4, 5, 6):
            summary.cell(row=row, column=column).number_format = "#,##0"
        summary[f"G{row}"].number_format = "0.0%"
    style_status_column(summary, 10, summary.max_row, 8)
    set_widths(summary, [10, 28, 12, 14, 12, 14, 14, 16])

    summary.print_area = f"A1:H{summary.max_row}"
    summary.page_setup.orientation = "landscape"
    summary.page_setup.fitToWidth = 1

    style_title(stores_sheet, 9, "Detalle de tiendas", f"Ordenado de mayor a menor avance · Región {region} · Corte {cut}")
    stores_headers = ["Ranking", "CeCo", "Tienda", "DM", "Realizadas", "Total", "Pendientes", "% Avance", "Estado"]
    stores_sheet.append(stores_headers)
    stores = sorted(data.get("stores", []), key=lambda item: (-item.get("compliance", 0), item.get("store", "")))
    for rank, item in enumerate(stores, 1):
        row = stores_sheet.max_row + 1
        stores_sheet.append([rank, item.get("ceco"), item.get("store"), item.get("dm"), item.get("completed", 0), item.get("expected", 0), item.get("expected", 0) - item.get("completed", 0), f"=IFERROR(E{row}/F{row},0)", status(item.get("compliance", 0))])
    style_header(stores_sheet, 4, 1, 9)
    style_table(stores_sheet, 5, stores_sheet.max_row, 9)
    stores_sheet.freeze_panes = "A5"
    stores_sheet.auto_filter.ref = f"A4:I{stores_sheet.max_row}"
    stores_sheet.conditional_formatting.add(f"H5:H{stores_sheet.max_row}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=GREEN, showValue=True))
    for row in range(5, stores_sheet.max_row + 1):
        stores_sheet[f"B{row}"].number_format = "@"
        for column in (1, 5, 6, 7):
            stores_sheet.cell(row=row, column=column).number_format = "#,##0"
        stores_sheet[f"H{row}"].number_format = "0.0%"
    style_status_column(stores_sheet, 5, stores_sheet.max_row, 9)
    set_widths(stores_sheet, [10, 12, 30, 32, 14, 12, 14, 14, 16])
    stores_sheet.print_title_rows = "1:4"
    stores_sheet.page_setup.orientation = "landscape"
    stores_sheet.page_setup.fitToWidth = 1

    style_title(activities_sheet, 7, "Avance por actividad", f"Lectura operativa · Región {region} · Corte {cut}")
    activity_headers = ["Orden", "Actividad", "Realizadas", "Pendientes", "Total", "% Avance", "Fecha compromiso"]
    activities_sheet.append(activity_headers)
    activities = sorted(data.get("activities", []), key=lambda item: (item.get("order", 999), item.get("name", "")))
    for item in activities:
        row = activities_sheet.max_row + 1
        activities_sheet.append([item.get("order"), item.get("name"), item.get("completedStores", 0), item.get("pendingStores", 0), item.get("completedStores", 0) + item.get("pendingStores", 0), f"=IFERROR(C{row}/E{row},0)", item.get("commitmentDateDisplay", "Sin fecha")])
    style_header(activities_sheet, 4, 1, 7)
    style_table(activities_sheet, 5, activities_sheet.max_row, 7)
    activities_sheet.freeze_panes = "A5"
    activities_sheet.auto_filter.ref = f"A4:G{activities_sheet.max_row}"
    activities_sheet.conditional_formatting.add(f"F5:F{activities_sheet.max_row}", DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1, color=GREEN, showValue=True))
    for row in range(5, activities_sheet.max_row + 1):
        for column in (1, 3, 4, 5):
            activities_sheet.cell(row=row, column=column).number_format = "#,##0"
        activities_sheet[f"F{row}"].number_format = "0.0%"
    set_widths(activities_sheet, [10, 42, 14, 14, 12, 14, 20])

    for index, ws in enumerate(workbook.worksheets):
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.sheet_properties.tabColor = [GREEN, DARK, GOOD][index]
        ws.sheet_view.zoomScale = 90
    return workbook


def main() -> None:
    parser = argparse.ArgumentParser(description="Genera un XLSX ejecutivo desde data/dashboard.json.")
    parser.add_argument("--data", type=Path, default=DEFAULT_DATA)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    data = json.loads(args.data.read_text(encoding="utf-8"))
    workbook = build_workbook(data)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.output)
    print(f"Excel generado: {args.output.relative_to(ROOT) if args.output.is_relative_to(ROOT) else args.output}")
    print(f"Resumen: {data['summary']['dms']} DM · {data['summary']['stores']} tiendas · {data['summary']['compliance']}% regional")


if __name__ == "__main__":
    main()
