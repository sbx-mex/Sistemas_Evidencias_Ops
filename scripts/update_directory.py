#!/usr/bin/env python3
"""Actualiza el Directorio multirregión y sincroniza los DM del CMS."""

from __future__ import annotations

import argparse
import os
import shutil
import tempfile
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from .build_dashboard import clean_text, find_directory_header, key_text, normalize_ceco, normalize_dm, short_dm_name, validate_xlsx
except ImportError:
    from build_dashboard import clean_text, find_directory_header, key_text, normalize_ceco, normalize_dm, short_dm_name, validate_xlsx

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_DIRECTORY = ROOT / "cms" / "Directorio.xlsx"
DEFAULT_CMS = ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx"

ORGANIZATION = [
    (1, "Centro's", "Raúl Sinohe Sierra Santa Maria", "Director Starbucks México", "assets/director/raul-sierra.webp", "Si", 1),
    (2, "Centro Centro", "Oliver Roberto Perez Briones", "Director Regional", "assets/director/oliver-perez.webp", "Si", 2),
    (2, "Centro Poniente", "Jorge Farrera Pinal", "Director Regional", "assets/director/jorge-farrera.webp", "Si", 3),
    (2, "Centro Sur", "Cielo Aide Morera Urrego", "Director Regional", "assets/director/cielo-morera.webp", "Si", 4),
    (2, "Centro Norte", "Jorge Antonio Alcantar Aguiar", "Director Regional", "assets/director/jorge-alcantar.webp", "Si", 5),
]


def atomic_copy(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile(prefix=f".{target.stem}-", suffix=target.suffix, dir=target.parent, delete=False) as handle:
        temporary = Path(handle.name)
    try:
        shutil.copy2(source, temporary)
        os.replace(temporary, target)
    finally:
        temporary.unlink(missing_ok=True)


def directory_people(path: Path) -> tuple[list[dict[str, str]], int, int, list[str], list[str]]:
    validate_xlsx(path, "el nuevo Directorio")
    workbook = load_workbook(path, read_only=True, data_only=True)
    candidates = []
    for ws in workbook.worksheets:
        try:
            header_row, headers = find_directory_header(ws)
        except ValueError:
            continue
        cols = {key_text(value): index for index, value in enumerate(headers) if clean_text(value)}
        if {"cc", "cc nombre", "region", "estatus", "dm"}.issubset(cols):
            candidates.append((ws.max_row, ws, header_row, cols))
    if not candidates:
        raise ValueError("El archivo no contiene CC, CC Nombre, Región, Estatus y DM")
    _, ws, header_row, cols = max(candidates, key=lambda item: item[0])

    stores: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        ceco = normalize_ceco(row[cols["cc"]])
        if not ceco:
            continue
        if ceco in seen:
            raise ValueError(f"CeCo duplicado en el nuevo Directorio: {ceco}")
        seen.add(ceco)
        stores.append({
            "ceco": ceco,
            "store": clean_text(row[cols["cc nombre"]]),
            "region": clean_text(row[cols["region"]]),
            "status": clean_text(row[cols["estatus"]]),
            "dm": normalize_dm(row[cols["dm"]]),
        })
    if not stores:
        raise ValueError("El nuevo Directorio no contiene tiendas")
    regions = sorted({item["region"] for item in stores if item["region"]}, key=key_text)
    open_stores = [item for item in stores if key_text(item["status"]) == "abierta"]
    missing_dm = [item["ceco"] for item in open_stores if item["dm"] == "DM pendiente"]
    return stores, len(open_stores), len(stores) - len(open_stores), regions, missing_dm


def existing_profiles(workbook) -> dict[str, dict[str, str]]:
    ws = workbook["Gerentes"]
    header_row = next(
        row for row in range(1, min(ws.max_row, 12) + 1)
        if key_text(ws.cell(row, 1).value) == "dm"
    )
    headers = {key_text(ws.cell(header_row, col).value): col for col in range(1, ws.max_column + 1)}
    profiles = {}
    for row in range(header_row + 1, ws.max_row + 1):
        dm = clean_text(ws.cell(row, headers["dm"]).value)
        if dm == "DM pendiente":
            continue
        profiles[key_text(dm)] = {
            "short": clean_text(ws.cell(row, headers.get("nombre corto", 2)).value) or short_dm_name(dm),
            "photo": clean_text(ws.cell(row, headers.get("foto webp", 3)).value),
        }
    return profiles


def sync_cms(cms_path: Path, stores: list[dict[str, str]]) -> tuple[int, int]:
    workbook = load_workbook(cms_path)
    profiles = existing_profiles(workbook)
    managers: dict[str, dict[str, str]] = {}
    for store in stores:
        if key_text(store["status"]) != "abierta":
            continue
        dm = store["dm"]
        if dm == "DM pendiente":
            continue
        item = managers.setdefault(key_text(dm), {"dm": dm, "regions": set()})
        item["regions"].add(store["region"])

    ws = workbook["Gerentes"]
    for merged in list(ws.merged_cells.ranges):
        if merged.min_row <= 2:
            ws.unmerge_cells(str(merged))
    ws.merge_cells("A1:F1")
    ws.merge_cells("A2:F2")
    ws["A1"] = "CMS · Gerentes de Distrito"
    ws["A2"] = "Nombre corto: primer nombre + primer apellido. El nombre completo permanece en DM; Foto WebP vacía indica pendiente."
    ws["A4"] = "DM"
    ws["B4"] = "Región"
    ws["C4"] = "Nombre corto"
    ws["D4"] = "Foto WebP"
    ws["E4"] = "Estado foto"
    ws["F4"] = "Activo"
    if ws.max_row > 4:
        ws.delete_rows(5, ws.max_row - 4)

    ordered = sorted(managers.values(), key=lambda item: (key_text(sorted(item["regions"], key=key_text)[0]), key_text(item["dm"])))
    for row_number, item in enumerate(ordered, 5):
        profile = profiles.get(key_text(item["dm"]), {})
        photo = profile.get("photo", "")
        ws.append([
            item["dm"], " · ".join(sorted(item["regions"], key=key_text)),
            profile.get("short") or short_dm_name(item["dm"]), photo,
            "Disponible" if photo else "Pendiente", "Si",
        ])
        for cell in ws[row_number]:
            cell.fill = PatternFill("solid", fgColor="EAF4EF")
            cell.font = Font(name="Aptos", size=10, color="24443A")
            cell.alignment = Alignment(vertical="center")

    dark = PatternFill("solid", fgColor="183F35")
    green = PatternFill("solid", fgColor="006241")
    for cell in ws[1]:
        cell.fill = green
        cell.font = Font(name="Aptos Display", size=12, bold=True, color="FFFFFF")
    for cell in ws[4]:
        cell.fill = dark
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    ws["A2"].fill = PatternFill("solid", fgColor="E6F2ED")
    ws["A2"].font = Font(name="Aptos", size=10, italic=True, color="36574D")
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = f"A4:F{4 + len(ordered)}"
    widths = {"A": 38, "B": 20, "C": 28, "D": 34, "E": 16, "F": 12}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width
    ws.data_validations.dataValidation = []
    active_validation = DataValidation(type="list", formula1='"Si,No"', allow_blank=False)
    photo_validation = DataValidation(type="list", formula1='"Disponible,Pendiente"', allow_blank=False)
    ws.add_data_validation(active_validation)
    ws.add_data_validation(photo_validation)
    active_validation.add(f"F5:F{max(104, 4 + len(ordered))}")
    photo_validation.add(f"E5:E{max(104, 4 + len(ordered))}")
    ws.conditional_formatting.add(
        f"E5:E{4 + len(ordered)}",
        FormulaRule(formula=["E5=\"Pendiente\""], fill=PatternFill("solid", fgColor="FFF1D6")),
    )

    if "Tiendas Abiertas" in workbook.sheetnames:
        del workbook["Tiendas Abiertas"]
    stores_ws = workbook.create_sheet("Tiendas Abiertas", 2)
    stores_ws.merge_cells("A1:E1")
    stores_ws.merge_cells("A2:E2")
    stores_ws["A1"] = "CMS · Tiendas Abiertas"
    stores_ws["A2"] = "Vista generada desde Directorio.xlsx. Sólo Estatus = Abierta alimenta tiendas, conteos y avance."
    stores_ws.append([])
    stores_ws.append(["CC", "CC Nombre", "Región", "Estatus", "DM"])
    open_stores = sorted(
        (item for item in stores if key_text(item["status"]) == "abierta"),
        key=lambda item: (key_text(item["region"]), key_text(item["dm"]), key_text(item["store"])),
    )
    for item in open_stores:
        stores_ws.append([item["ceco"], item["store"], item["region"], item["status"], item["dm"]])
    for cell in stores_ws[1]:
        cell.fill = green
        cell.font = Font(name="Aptos Display", size=12, bold=True, color="FFFFFF")
    stores_ws["A2"].fill = PatternFill("solid", fgColor="E6F2ED")
    stores_ws["A2"].font = Font(name="Aptos", size=10, italic=True, color="36574D")
    for cell in stores_ws[4]:
        cell.fill = dark
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    for row in stores_ws.iter_rows(min_row=5, max_row=4 + len(open_stores), min_col=1, max_col=5):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="F4F8F6")
            cell.font = Font(name="Aptos", size=10, color="24443A")
            cell.alignment = Alignment(vertical="center")
    stores_ws.freeze_panes = "A5"
    stores_ws.auto_filter.ref = f"A4:E{4 + len(open_stores)}"
    for column, width in {"A": 12, "B": 34, "C": 22, "D": 16, "E": 40}.items():
        stores_ws.column_dimensions[column].width = width
    stores_ws.sheet_view.showGridLines = False

    if "Organigrama" not in workbook.sheetnames:
        org_ws = workbook.create_sheet("Organigrama", 2)
        org_ws.merge_cells("A1:G1")
        org_ws.merge_cells("A2:G2")
        org_ws["A1"] = "CMS · Organigrama Región | Centro's"
        org_ws["A2"] = "Edita nombre, fotografía, estado y orden. La web muestra sólo los cuatro RD activos; el rol no se repite en la tarjeta."
        org_ws.append([])
        org_ws.append(["Nivel", "Región", "Nombre", "Rol", "Foto WebP", "Activo", "Orden"])
        for item in ORGANIZATION:
            org_ws.append(list(item))
    else:
        org_ws = workbook["Organigrama"]
    for cell in org_ws[1]:
        cell.fill = green
        cell.font = Font(name="Aptos Display", size=12, bold=True, color="FFFFFF")
    org_ws["A2"].fill = PatternFill("solid", fgColor="FFF2BF")
    org_ws["A2"].font = Font(name="Aptos", size=10, italic=True, color="5C4A00")
    for cell in org_ws[4]:
        cell.fill = dark
        cell.font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    for row in org_ws.iter_rows(min_row=5, max_row=org_ws.max_row, min_col=1, max_col=7):
        for cell in row:
            cell.fill = PatternFill("solid", fgColor="F4F8F6")
            cell.font = Font(name="Aptos", size=10, color="24443A")
            cell.alignment = Alignment(vertical="center")
    org_ws.freeze_panes = "A5"
    org_ws.auto_filter.ref = f"A4:G{org_ws.max_row}"
    for column, width in {"A": 10, "B": 22, "C": 36, "D": 27, "E": 38, "F": 12, "G": 10}.items():
        org_ws.column_dimensions[column].width = width
    org_ws.sheet_view.showGridLines = False
    org_ws.data_validations.dataValidation = []
    org_active = DataValidation(type="list", formula1='"Si,No"', allow_blank=False)
    org_ws.add_data_validation(org_active)
    org_active.add("F5:F100")

    config = workbook["Configuracion"]
    config_header = next(row for row in range(1, min(config.max_row, 12) + 1) if key_text(config.cell(row, 1).value) == "clave")
    config_rows = {clean_text(config.cell(row, 1).value): row for row in range(config_header + 1, config.max_row + 1)}
    changes = {
        "region": ("Todas", "Incluir todas las regiones presentes en Directorio.xlsx"),
        "directorySheet": ("Directorio", "Hoja preferida; Python detecta otra hoja válida si cambia el nombre"),
        "onlyOpenStores": ("Si", "Aplicar el filtro de estatus definido en includedStoreStatuses"),
        "includedStoreStatuses": ("Abierta", "Único estatus incluido en tiendas, conteos y avance"),
    }
    for key, (value, description) in changes.items():
        row = config_rows.get(key)
        if row is None:
            row = config.max_row + 1
            config.cell(row, 1).value = key
            config_rows[key] = row
            for column in range(1, 4):
                source = config.cell(row - 1, column)
                target = config.cell(row, column)
                if source.has_style:
                    target._style = copy(source._style)
                target.alignment = copy(source.alignment)
        config.cell(row, 2).value = value
        config.cell(row, 3).value = description

    with tempfile.NamedTemporaryFile(prefix=f".{cms_path.stem}-", suffix=".xlsx", dir=cms_path.parent, delete=False) as handle:
        temporary = Path(handle.name)
    try:
        workbook.save(temporary)
        os.replace(temporary, cms_path)
    finally:
        temporary.unlink(missing_ok=True)
    pending = sum(not profiles.get(key, {}).get("photo") for key in managers)
    return len(managers), pending


def main() -> None:
    parser = argparse.ArgumentParser(description="Actualiza Directorio.xlsx y agrupa regiones/DM en el CMS")
    parser.add_argument("source", type=Path, help="Nuevo Directorio.xlsx")
    parser.add_argument("--directory", type=Path, default=DEFAULT_DIRECTORY)
    parser.add_argument("--cms", type=Path, default=DEFAULT_CMS)
    args = parser.parse_args()
    stores, open_count, excluded_count, regions, missing_dm = directory_people(args.source)
    atomic_copy(args.source, args.directory)
    managers, pending_photos = sync_cms(args.cms, stores)
    print(f"Directorio actualizado · {open_count} tiendas Abierta · {excluded_count} excluidas · {len(regions)} regiones · {managers} DM")
    print(f"Fotos DM · {managers - pending_photos} disponibles · {pending_photos} pendientes")
    if missing_dm:
        print("Asignación DM pendiente · " + ", ".join(missing_dm))


if __name__ == "__main__":
    main()
