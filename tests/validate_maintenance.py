#!/usr/bin/env python3
"""Prueba mantenimiento seguro, CMS flexible y persistencia atómica."""

from __future__ import annotations

import shutil
import sys
import tempfile
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.build_dashboard import build_payload, find_header, key_text, load_cms, validate_webp_asset
from scripts.build_dashboard import (
    STABILITY_CONTROLS, ensure_source_stability, find_directory_header,
    load_directory, load_settings, normalize_allowed_hosts, source_fingerprints,
)
from scripts.clean_obsolete import existing_obsolete_files
from scripts.io_utils import atomic_output_path, atomic_write_text
from scripts.validate_sources import validate_cms_engine


def test_atomic_writes(temp: Path) -> None:
    target = temp / "resultado.json"
    target.write_text("estable", encoding="utf-8")
    with atomic_output_path(target) as temporary:
        temporary.write_text("actualizado", encoding="utf-8")
    assert target.read_text(encoding="utf-8") == "actualizado"

    try:
        with atomic_output_path(target) as temporary:
            temporary.write_text("incompleto", encoding="utf-8")
            raise RuntimeError("fallo simulado")
    except RuntimeError:
        pass
    else:
        raise AssertionError("El fallo simulado no se propagó")
    assert target.read_text(encoding="utf-8") == "actualizado"
    assert not list(temp.glob(".*.tmp"))

    atomic_write_text(target, "final\n")
    assert target.read_text(encoding="utf-8") == "final\n"


def test_obsolete_detection(temp: Path) -> None:
    project = temp / "limpieza"
    for directory in (project / "data", project / "exports", project / "cms", project / "scripts" / "__pycache__"):
        directory.mkdir(parents=True, exist_ok=True)
    (project / "data" / ".dashboard.json.abc.tmp").write_text("parcial", encoding="utf-8")
    (project / "exports" / "Resumen.pdf.tmp").write_text("parcial", encoding="utf-8")
    (project / "cms" / "~$Sistema.xlsx").write_text("bloqueo", encoding="utf-8")
    (project / "scripts" / "__pycache__" / "motor.pyc").write_bytes(b"cache")
    assert existing_obsolete_files(project) == [
        "cms/~$Sistema.xlsx",
        "data/.dashboard.json.abc.tmp",
        "exports/Resumen.pdf.tmp",
        "scripts/__pycache__/motor.pyc",
    ]


def expect_error(action, text: str) -> None:
    try:
        action()
    except (RuntimeError, ValueError) as error:
        assert text in str(error)
    else:
        raise AssertionError(f"No se bloqueó el escenario inseguro: {text}")


def test_source_and_host_guards(temp: Path) -> None:
    source = temp / "fuente.xlsx"
    source.write_bytes(b"version-1")
    paths = {"Forms": source}
    before = source_fingerprints(paths)
    source.write_bytes(b"version-2")
    expect_error(lambda: ensure_source_stability(before, paths), "cambiaron durante")

    assert normalize_allowed_hosts("GRUPOVIPS-my.sharepoint.com.") == {
        "grupovips-my.sharepoint.com"
    }
    expect_error(lambda: normalize_allowed_hosts("https://example.com/ruta"), "Dominio")
    expect_error(lambda: normalize_allowed_hosts(""), "al menos un dominio")


def test_structural_guards(temp: Path) -> None:
    original_cms = ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx"

    duplicate_config = temp / "cms_config_duplicada.xlsx"
    shutil.copy2(original_cms, duplicate_config)
    workbook = load_workbook(duplicate_config)
    sheet = workbook["Configuracion"]
    header_row, cols = find_header(sheet, {"clave", "valor"})
    row = sheet.max_row + 1
    sheet.cell(row, cols["clave"] + 1, "PROJECTNAME")
    sheet.cell(row, cols["valor"] + 1, "Duplicado")
    workbook.save(duplicate_config)
    expect_error(lambda: load_cms(duplicate_config), "configuración duplicada")

    duplicate_activity = temp / "cms_actividad_duplicada.xlsx"
    shutil.copy2(original_cms, duplicate_activity)
    workbook = load_workbook(duplicate_activity)
    sheet = workbook["Actividades"]
    header_row, cols = find_header(sheet, {
        "orden", "actividad", "descripcion", "fecha inicio", "fecha limite", "activo",
    })
    first = header_row + 1
    target = sheet.max_row + 1
    for column in range(1, sheet.max_column + 1):
        sheet.cell(target, column, sheet.cell(first, column).value)
    sheet.cell(target, cols["orden"] + 1, 999)
    workbook.save(duplicate_activity)
    expect_error(lambda: load_cms(duplicate_activity), "duplicadas")

    no_active = temp / "cms_sin_activas.xlsx"
    shutil.copy2(original_cms, no_active)
    workbook = load_workbook(no_active)
    sheet = workbook["Actividades"]
    header_row, cols = find_header(sheet, {
        "orden", "actividad", "descripcion", "fecha inicio", "fecha limite", "activo",
    })
    for row in range(header_row + 1, sheet.max_row + 1):
        if sheet.cell(row, cols["actividad"] + 1).value:
            sheet.cell(row, cols["activo"] + 1, "No")
    workbook.save(no_active)
    expect_error(lambda: load_cms(no_active), "no contiene actividades activas")

    duplicate_directory = temp / "directorio_duplicado.xlsx"
    shutil.copy2(ROOT / "cms" / "Directorio.xlsx", duplicate_directory)
    _, _, cms_settings, _ = load_cms(original_cms)
    settings = load_settings(ROOT / "config" / "settings.json", cms_settings)
    workbook = load_workbook(duplicate_directory)
    sheet = workbook[settings["directorySheet"]]
    header_row, _ = find_directory_header(sheet)
    first = header_row + 1
    target = sheet.max_row + 1
    for column in range(1, sheet.max_column + 1):
        sheet.cell(target, column, sheet.cell(first, column).value)
    workbook.save(duplicate_directory)
    expect_error(lambda: load_directory(duplicate_directory, settings), "CeCo duplicado")


def test_flexible_cms(temp: Path) -> None:
    baseline_activities, _, _, _ = load_cms(
        ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx"
    )
    expected_count = len(baseline_activities)
    cms = temp / "cms_editado.xlsx"
    shutil.copy2(ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx", cms)
    workbook = load_workbook(cms)

    activities = workbook["Actividades"]
    header_row, cols = find_header(activities, {
        "orden", "actividad", "descripcion", "fecha inicio", "fecha limite",
        "activo", "evidencia requerida", "prioridad", "estado fecha",
    })
    first = header_row + 1
    activities["A1"] = "CMS actualizado por operación"
    activities["A2"] = "Las notas y celdas auxiliares pueden cambiar sin romper el motor."
    activities.cell(first, cols["descripcion"] + 1, "Descripción actualizada desde CMS")
    activities.cell(first, cols["orden"] + 1, "por definir")
    activities.cell(first, cols["fecha inicio"] + 1, "fecha en revisión")
    activities.cell(first, cols["evidencia requerida"] + 1, "En revisión")
    activities.cell(first, cols["prioridad"] + 1, "Urgente")
    activities.cell(first, cols["estado fecha"] + 1, "Estado informativo manual")

    last_activity_row = max(
        row for row in range(header_row + 1, activities.max_row + 1)
        if activities.cell(row, cols["actividad"] + 1).value
    )
    draft = last_activity_row + 1
    activities.cell(draft, cols["orden"] + 1, "pendiente")
    activities.cell(draft, cols["actividad"] + 1, "Borrador sin publicar")
    activities.cell(draft, cols["fecha inicio"] + 1, "por definir")
    activities.cell(draft, cols["fecha limite"] + 1, "por definir")
    activities.cell(draft, cols["activo"] + 1, "")

    config = workbook["Configuracion"]
    config_header, config_cols = find_header(config, {"clave", "valor"})
    for row in range(config_header + 1, config.max_row + 1):
        key = config.cell(row, config_cols["clave"] + 1).value
        if key == "projectName":
            config.cell(row, config_cols["valor"] + 1, "")
        elif key == "requireEvidence":
            config.cell(row, config_cols["valor"] + 1, "En revisión")
    workbook.save(cms)

    audit = validate_cms_engine(cms)
    assert audit["activities"] == expected_count
    assert audit["drafts"] >= 1
    assert audit["fallbackOrders"] == 1
    assert audit["manualStatuses"] == 1
    assert audit["customPriorities"] == 1
    assert audit["invalidDates"] == 1
    assert audit["safeDefaults"] >= 3

    loaded, _, cms_settings, _ = load_cms(cms)
    assert len(loaded) == expected_count
    assert "Borrador sin publicar" not in {item["name"] for item in loaded}
    edited = next(item for item in loaded if item["name"] == "Roll Out")
    assert edited["description"] == "Descripción actualizada desde CMS"
    assert edited["priority"] == "Urgente"
    assert edited["requireEvidence"] is True
    assert edited["startDate"] is None
    assert "projectName" not in cms_settings and "requireEvidence" not in cms_settings

    payload = build_payload(
        ROOT / "cms" / "Sistema de Evidencias OPS.xlsx",
        ROOT / "cms" / "Directorio.xlsx",
        ROOT / "config" / "settings.json",
        cms,
    )
    assert payload["project"] == "Sistema de Evidencias OPS"
    assert payload["summary"]["activities"] == expected_count
    assert tuple(payload["quality"]["stabilityControls"]) == STABILITY_CONTROLS
    assert all(payload["quality"]["stabilityControls"].values())


def test_activity_row_count_independence(temp: Path) -> None:
    """Agregar o borrar filas válidas no cambia el contrato basado en encabezados."""
    original = ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx"
    baseline, _, _, _ = load_cms(original)
    baseline_names = {item["name"] for item in baseline}

    reduced = temp / "cms_una_fila_menos.xlsx"
    shutil.copy2(original, reduced)
    workbook = load_workbook(reduced)
    sheet = workbook["Actividades"]
    header_row, cols = find_header(sheet, {
        "orden", "actividad", "descripcion", "fecha inicio", "fecha limite", "activo",
    })
    removable = next(
        row for row in range(header_row + 1, sheet.max_row + 1)
        if sheet.cell(row, cols["actividad"] + 1).value
        and str(sheet.cell(row, cols["activo"] + 1).value).strip().casefold() in {"si", "sí"}
    )
    removed_name = str(sheet.cell(removable, cols["actividad"] + 1).value).strip()
    sheet.delete_rows(removable, 1)
    workbook.save(reduced)
    reduced_activities, _, _, _ = load_cms(reduced)
    assert len(reduced_activities) == len(baseline) - 1
    assert {item["name"] for item in reduced_activities} == baseline_names - {removed_name}

    expanded = temp / "cms_una_fila_mas.xlsx"
    shutil.copy2(original, expanded)
    workbook = load_workbook(expanded)
    sheet = workbook["Actividades"]
    header_row, cols = find_header(sheet, {
        "orden", "actividad", "descripcion", "fecha inicio", "fecha limite", "activo",
    })
    target = sheet.max_row + 1
    sheet.cell(target, cols["orden"] + 1, 999)
    sheet.cell(target, cols["actividad"] + 1, "Actividad de estabilidad")
    sheet.cell(target, cols["descripcion"] + 1, "Prueba automática")
    sheet.cell(target, cols["activo"] + 1, "Sí")
    workbook.save(expanded)
    expanded_activities, _, _, _ = load_cms(expanded)
    assert len(expanded_activities) == len(baseline) + 1
    assert {item["name"] for item in expanded_activities} == baseline_names | {"Actividad de estabilidad"}


def test_manager_photo_guards(temp: Path) -> None:
    """Detecta el WebP canónico y bloquea rutas o contenidos inseguros."""
    cms = temp / "cms_foto_autodetectada.xlsx"
    shutil.copy2(ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx", cms)
    workbook = load_workbook(cms)
    sheet = workbook["Gerentes"]
    header_row, cols = find_header(sheet, {"dm", "nombre corto", "foto webp", "activo"})
    target_dm = "Adriana Alejandra Tanus Buhler"
    target_row = next(
        row for row in range(header_row + 1, sheet.max_row + 1)
        if key_text(sheet.cell(row, cols["dm"] + 1).value) == key_text(target_dm)
    )
    sheet.cell(target_row, cols["foto webp"] + 1).value = None
    workbook.save(cms)

    _, managers, _, _ = load_cms(cms)
    profile = managers[key_text(target_dm)]
    assert profile["photo"] == "assets/dm/adriana-tanus.webp"
    assert profile["photoSource"] == "Detectada"

    invalid_root = ROOT / "assets" / ".photo-validation"
    invalid_root.mkdir(exist_ok=True)
    invalid_photo = invalid_root / "contenido-falso.webp"
    try:
        invalid_photo.write_text("esto no es una imagen", encoding="utf-8")
        relative = invalid_photo.relative_to(ROOT).as_posix()
        expect_error(lambda: validate_webp_asset(relative, "prueba"), "no es un WebP válido")
        expect_error(lambda: validate_webp_asset("../fuera.webp", "prueba"), "sale del proyecto")
        expect_error(lambda: validate_webp_asset("assets/dm/adriana-tanus.jpg", "prueba"), "Ruta WebP inválida")
        expect_error(lambda: validate_webp_asset("assets/dm/no-existe.webp", "prueba"), "No existe")
    finally:
        invalid_photo.unlink(missing_ok=True)
        invalid_root.rmdir()


def main() -> None:
    with tempfile.TemporaryDirectory() as temp_dir:
        temp = Path(temp_dir)
        test_atomic_writes(temp)
        test_obsolete_detection(temp)
        test_source_and_host_guards(temp)
        test_structural_guards(temp)
        test_flexible_cms(temp)
        test_activity_row_count_independence(temp)
        test_manager_photo_guards(temp)
    print("Mantenimiento aprobado · CMS flexible · escritura atómica · borradores seguros")


if __name__ == "__main__":
    main()
