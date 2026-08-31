#!/usr/bin/env python3
"""Prueba cambios de columnas/filas del Excel exportado por Microsoft Forms."""

from __future__ import annotations

import sys
import tempfile
import shutil
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

from scripts.build_dashboard import boolean_answer, build_payload, clean_text, evidence_header_activity, load_cms, load_responses


BASE = ["Id", "Hora de inicio", "Hora de finalización", "Correo electrónico", "Nombre"]
ACTIVITY = "Selecciona la actividad que deseas registrar"


def save_book(path: Path, headers: list[str], rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def timestamps(index: int) -> tuple[datetime, datetime]:
    start = datetime(2026, 8, 29, 10, 0) + timedelta(minutes=index)
    return start, start + timedelta(seconds=30)


def main() -> None:
    allowed = "https://grupovips-my.sharepoint.com/evidencias"
    with tempfile.TemporaryDirectory() as temp_dir:
        temp = Path(temp_dir)

        # Escenario 1: exportación ancha, columnas reordenadas y una por actividad.
        wide = temp / "wide.xlsx"
        start, finish = timestamps(1)
        wide_headers = BASE + ["CeCo", ACTIVITY, "¿Comentario operativo?", "Evidencia_Lay_Out", "Evidencia_RollOut"]
        save_book(wide, wide_headers, [[1, start, finish, "", "Prueba", "38115", "Roll Out", "x", "", f"{allowed}/rollout.jpg"]])
        rows, schema = load_responses(wide)
        assert rows[0]["evidence"].endswith("rollout.jpg")
        assert rows[0]["evidenceSourceHeader"] == "Evidencia_RollOut"
        assert rows[0]["confirmed"] is True and schema["confirmationHeaders"] == []

        # Escenario 2: formato largo; Forms agrega nuevas respuestas hacia abajo.
        long_book = temp / "long.xlsx"
        long_headers = BASE + ["CeCo", ACTIVITY, "¿Confirmas que realizaste la actividad seleccionada?", "Evidencia del avance"]
        start1, finish1 = timestamps(2)
        start2, finish2 = timestamps(3)
        save_book(long_book, long_headers, [
            [2, start1, finish1, "", "Prueba", "38115", "Roll Out", "No", f"{allowed}/uno.jpg"],
            [3, start2, finish2, "", "Prueba", "38149", "Lay Out", "Sí", f"{allowed}/dos.jpg"],
        ])
        rows, schema = load_responses(long_book)
        assert len(rows) == 2 and all(row["confirmed"] for row in rows)
        assert all(row["explicitNo"] is False and row["confirmedAnswer"] == "Sí" for row in rows)
        assert schema["evidenceHeaderMap"] == {"Evidencia del avance": "generic"}

        # El motor tolera texto UTF-8 que una exportación intermedia haya leído
        # accidentalmente como Latin-1, pero el repositorio conserva UTF-8 limpio.
        mojibake_yes = "Sí".encode("utf-8").decode("latin-1")
        assert clean_text(mojibake_yes) == "Sí"
        assert boolean_answer(mojibake_yes) is True

        # Escenario 3: encabezados base duplicados; se toma el único valor poblado.
        duplicate = temp / "duplicate.xlsx"
        start, finish = timestamps(4)
        duplicate_headers = BASE + ["CeCo", "CeCo", ACTIVITY, ACTIVITY, "Evidencia_QR_Qualtrics"]
        save_book(duplicate, duplicate_headers, [[4, start, finish, "", "Prueba", "", "38965", "", "QR - Qualtrics", f"{allowed}/qr.jpg"]])
        rows, schema = load_responses(duplicate)
        assert rows[0]["ceco"] == "38965" and rows[0]["activity"] == "QR - Qualtrics"
        assert schema["rowConflicts"] == []

        # Escenario 4: dos evidencias incompatibles no se mezclan ni se adivinan.
        ambiguous = temp / "ambiguous.xlsx"
        start, finish = timestamps(5)
        ambiguous_headers = BASE + ["CeCo", ACTIVITY, "Evidencia_QR_Qualtrics", "Evidencia_Lay_Out"]
        save_book(ambiguous, ambiguous_headers, [[5, start, finish, "", "Prueba", "38115", "Roll Out", f"{allowed}/qr.jpg", f"{allowed}/layout.jpg"]])
        rows, schema = load_responses(ambiguous)
        assert rows[0]["evidence"] == "" and rows[0]["confirmed"] is True
        assert schema["evidenceIssues"]["ambiguous-evidence"] == [2]

        # Escenario 5: una sola evidencia en la columna incorrecta tampoco se reasigna.
        mismatched = temp / "mismatched.xlsx"
        start, finish = timestamps(6)
        mismatched_headers = BASE + ["CeCo", ACTIVITY, "Evidencia_QR_Qualtrics"]
        save_book(mismatched, mismatched_headers, [[6, start, finish, "", "Prueba", "38115", "Roll Out", f"{allowed}/qr.jpg"]])
        rows, schema = load_responses(mismatched)
        assert rows[0]["evidence"] == "" and rows[0]["confirmed"] is True
        assert schema["evidenceIssues"]["mismatched-evidence-column"] == [2]

        # Escenario 6: dos actividades fuera del catálogo activo CMS se ignoran.
        # Una no existe y otra sí existe en el CMS, pero está desactivada. El orden
        # de columnas y las filas repetidas no deben afectar avance ni fecha de corte.
        hidden = temp / "hidden.xlsx"
        hidden_cms = temp / "cms-with-inactive.xlsx"
        shutil.copy2(ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx", hidden_cms)
        cms_book = load_workbook(hidden_cms)
        activity_sheet = cms_book["Actividades"]
        cms_headers = {
            clean_text(cell.value): index
            for index, cell in enumerate(next(activity_sheet.iter_rows(min_row=4, max_row=4)), 1)
            if clean_text(cell.value)
        }
        for row_number in range(5, activity_sheet.max_row + 1):
            if clean_text(activity_sheet.cell(row_number, cms_headers["Actividad"]).value) == "Roll Out":
                activity_sheet.cell(row_number, cms_headers["Activo"]).value = "No"
                break
        cms_book.save(hidden_cms)

        start1, finish1 = timestamps(7)
        start2, finish2 = timestamps(8)
        start3, finish3 = timestamps(9)
        hidden_headers = BASE + ["Evidencia_RollOut", "CeCo", "Evidencia_Nueva_Actividad", ACTIVITY]
        save_book(hidden, hidden_headers, [
            [6, start1, finish1, "", "Prueba", f"{allowed}/rollout.jpg", "38115", "", "Roll Out"],
            [7, start2, finish2, "", "Prueba", "", "38115", f"{allowed}/nueva.jpg", "Nueva Actividad Forms"],
            [8, start3, finish3, "", "Prueba", "", "99999", f"{allowed}/repetida.jpg", "Nueva Actividad Forms"],
        ])
        payload = build_payload(
            hidden,
            ROOT / "cms" / "Centro Norte_Directorio.xlsx",
            ROOT / "config" / "settings.json",
            hidden_cms,
        )
        hidden_cms_activities, _, _, _ = load_cms(hidden_cms)
        assert payload["summary"]["activities"] == len(hidden_cms_activities)
        assert "Roll Out" not in {item["name"] for item in payload["activities"]}
        assert payload["submissions"] == []
        assert payload["quality"]["hiddenActivities"] == ["Nueva Actividad Forms", "Roll Out"]
        assert payload["quality"]["hiddenActivityRows"] == [2, 3, 4]
        assert payload["quality"]["unknownCeCos"] == []
        assert payload["summary"]["validResponses"] == 0
        assert payload["summary"]["completedCompletions"] == 0
        assert payload["lastUpdated"] is None and payload["lastUpdatedDisplay"] == "Sin respuestas"
        cms_activities, _, _, _ = load_cms(ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx")

        # Escenario 6b: Roll Out se enlaza con Evidencia_RollOut por nombre,
        # aunque cambie el orden. Dos respuestas del mismo par publican sólo la última.
        repeated = temp / "repeated-rollout.xlsx"
        old_start, old_finish = timestamps(11)
        new_start, new_finish = timestamps(12)
        repeated_headers = BASE + ["Evidencia_RollOut", ACTIVITY, "CeCo"]
        save_book(repeated, repeated_headers, [
            [11, old_start, old_finish, "", "Prueba", f"{allowed}/rollout-anterior.jpg", "Roll Out", "38115"],
            [12, new_start, new_finish, "", "Prueba", f"{allowed}/rollout-vigente.jpg", "Roll Out", "38115"],
        ])
        repeated_payload = build_payload(
            repeated,
            ROOT / "cms" / "Centro Norte_Directorio.xlsx",
            ROOT / "config" / "settings.json",
            ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx",
        )
        assert repeated_payload["summary"]["completedCompletions"] == 1
        assert repeated_payload["summary"]["validResponses"] == 1
        assert repeated_payload["quality"]["duplicateValidResponses"] == 1
        assert len(repeated_payload["submissions"]) == 1
        assert repeated_payload["submissions"][0]["evidenceUrl"].endswith("rollout-vigente.jpg")
        assert repeated_payload["quality"]["responseSchema"]["evidenceHeaderMap"]["Evidencia_RollOut"] == "rollout"
        assert repeated_payload["quality"]["stabilityScore"] == "10/10"
        assert len(repeated_payload["quality"]["stabilityControls"]) == 10

        # Escenario 7: portada previa, encabezado desplazado y campos personales ausentes.
        shifted = temp / "shifted.xlsx"
        workbook = Workbook()
        cover = workbook.active
        cover.title = "Instrucciones"
        cover["A1"] = "Exportación Microsoft Forms"
        data_sheet = workbook.create_sheet("Respuestas")
        data_sheet.append(["Sistema de Evidencias OPS"])
        data_sheet.append([])
        shifted_headers = ["CeCo", ACTIVITY, "Evidencia_Rack_FHW", "Nueva columna"]
        data_sheet.append(shifted_headers)
        data_sheet.append(["38590", "Rack FHW", f"{allowed}/rack.jpg", "futuro"])
        workbook.save(shifted)
        rows, schema = load_responses(shifted)
        assert schema["sheet"] == "Respuestas" and schema["headerRow"] == 3
        assert rows[0]["ceco"] == "38590" and rows[0]["confirmed"] is True
        assert rows[0]["email"] == "" and rows[0]["name"] == ""
        assert rows[0]["id"].startswith("respuesta-")

        # Escenario 8: el orden cambia y los encabezados pueden ser aproximados
        # o únicamente el nombre de la actividad. El CMS aporta el catálogo válido.
        flexible = temp / "flexible.xlsx"
        start1, finish1 = timestamps(8)
        start2, finish2 = timestamps(9)
        flexible_headers = [
            "Columna nueva", "Fotografia SM", "CeCo", ACTIVITY,
            "Evidencia_Programacion_Horno_Merry_Focaccia", "Hora de finalización",
        ]
        save_book(flexible, flexible_headers, [
            ["x", "", "38333", "Programacion Hornos Merry - Focaccia", f"{allowed}/horno.jpg", finish1],
            ["x", f"{allowed}/foto.jpg", "38115", "Fotografia - SM", "", finish2],
        ])
        activity_names = [
            "Roll Out", "Programacion Hornos Merry - Focaccia", "Rack FHW",
            "QR - Qualtrics", "Max & Min", "Fotografia - SM", "Mandil Verde", "Lay Out",
        ]
        rows, schema = load_responses(flexible, activity_names)
        assert [row["activity"] for row in rows] == [
            "Programacion Hornos Merry - Focaccia", "Fotografia - SM",
        ]
        assert rows[0]["evidence"].endswith("horno.jpg")
        assert rows[1]["evidence"].endswith("foto.jpg")
        assert schema["evidenceHeaderMatch"]["Evidencia_Programacion_Horno_Merry_Focaccia"] == "affinity"
        assert schema["evidenceHeaderMatch"]["Fotografia SM"] == "exact"
        assert schema["evidenceIssues"] == {}

        # Escenario 8b: mayúsculas inestables, acentos y un error menor conservan
        # la actividad correcta sin depender de escritura exacta.
        unstable = temp / "unstable_headers.xlsx"
        start, finish = timestamps(10)
        unstable_headers = BASE + ["CeCo", ACTIVITY, "Evidencia_ACtivacion_PSL_Sharpiee"]
        save_book(unstable, unstable_headers, [[
            10, start, finish, "", "Prueba", "38115", "Activacion PSL Sharpie", f"{allowed}/psl.jpg",
        ]])
        rows, schema = load_responses(unstable, ["Activacion PSL Sharpie", "Activacion PSL Verano"])
        assert rows[0]["evidence"].endswith("psl.jpg")
        assert rows[0]["evidenceSourceHeader"] == "Evidencia_ACtivacion_PSL_Sharpiee"
        assert schema["evidenceHeaderMap"]["Evidencia_ACtivacion_PSL_Sharpiee"] == "activacionpslsharpie"
        assert schema["evidenceHeaderMatch"]["Evidencia_ACtivacion_PSL_Sharpiee"] == "affinity"
        exact_key, exact_type = evidence_header_activity(
            "EVIDENCIA_ÁCTIVACION_PSL_SHARPIE", ["Activacion PSL Sharpie"]
        )
        assert exact_key == "activacionpslsharpie" and exact_type == "exact"

        # Escenario 9: cambiar el orden editorial del CMS sólo cambia la
        # presentación; el cumplimiento continúa cruzándose por nombre.
        reordered_cms = temp / "cms_reordered.xlsx"
        cms_book = load_workbook(ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx")
        activity_sheet = cms_book["Actividades"]
        for row_number in range(5, activity_sheet.max_row + 1):
            if activity_sheet.cell(row_number, 2).value and activity_sheet.cell(row_number, 1).value is not None:
                activity_sheet.cell(row_number, 1).value = 100 - int(activity_sheet.cell(row_number, 1).value)
        cms_book.save(reordered_cms)
        payload = build_payload(
            flexible,
            ROOT / "cms" / "Centro Norte_Directorio.xlsx",
            ROOT / "config" / "settings.json",
            reordered_cms,
        )
        assert payload["summary"]["completedCompletions"] == 2
        assert [item["focusRank"] for item in payload["activities"]] == list(range(1, len(payload["activities"]) + 1))
        pending_dates = [item["endDate"] for item in payload["activities"] if item["pendingStores"] and item["endDate"]]
        assert pending_dates == sorted(pending_dates)
        stores = {store["ceco"]: store for store in payload["stores"]}
        assert stores["38333"]["activities"]["Programacion Hornos Merry - Focaccia"] is True
        assert stores["38115"]["activities"]["Fotografia - SM"] is True

        # Escenario 10: preguntas Sí/No duplicadas, reordenadas y con texto adicional.
        # La respuesta se asocia a la actividad elegida en la misma fila.
        conditional = temp / "conditional.xlsx"
        conditional_headers = BASE + [
            "CeCo", ACTIVITY,
            "¿ Tienes Horno Merry Chef ?", "¿Tienes Horno Merry Chef? (2)",
            "¿Cuentas con Community Board?",
            "Evidencia_Programacion_Hornos_Merry_Focaccia", "Evidencia_Community_Board",
        ]
        enrique_cecos = ["38333", "38339", "38368", "38401", "38456", "38515", "38604", "38862", "38894", "43193"]
        conditional_rows = []
        for index, ceco in enumerate(enrique_cecos):
            start, finish = timestamps(20 + index)
            if index < 8:
                conditional_rows.append([
                    20 + index, start, finish, "", "Prueba", ceco,
                    "Programacion Hornos Merry - Focaccia", "Sí, contamos con horno", "", "",
                    f"{allowed}/horno-{index}.jpg", "",
                ])
            else:
                conditional_rows.append([
                    20 + index, start, finish, "", "Prueba", ceco,
                    "Programacion Hornos Merry - Focaccia", "No contamos con Horno Merry Chef", "No", "",
                    "", "",
                ])
        start, finish = timestamps(40)
        conditional_rows.append([
            40, start, finish, "", "Prueba", "38115", "Community Board", "", "", "Sí",
            "", f"{allowed}/community.jpg",
        ])
        start, finish = timestamps(41)
        conditional_rows.append([
            41, start, finish, "", "Prueba", "38119", "Community Board", "", "", "No aplica en esta tienda",
            "", "",
        ])
        save_book(conditional, conditional_headers, conditional_rows)
        payload = build_payload(
            conditional,
            ROOT / "cms" / "Centro Norte_Directorio.xlsx",
            ROOT / "config" / "settings.json",
            ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx",
        )
        activity_map = {item["name"]: item for item in payload["activities"]}
        horno = activity_map["Programacion Hornos Merry - Focaccia"]
        community = activity_map["Community Board"]
        assert horno["completedStores"] == 8 and horno["notApplicableStores"] == 2
        assert community["completedStores"] == 1 and community["notApplicableStores"] == 1
        enrique = next(item for item in payload["dms"] if item["dm"] == "Enrique Cesar Flores")
        expected_enrique = len(enrique_cecos) * len(cms_activities) - 2
        assert enrique["completed"] == 8 and enrique["expected"] == expected_enrique
        assert payload["summary"]["notApplicableCompletions"] == 3
        assert payload["quality"]["responseSchema"]["applicabilityIssues"] == {}
        assert len(payload["quality"]["responseSchema"]["applicabilityHeaders"]) == 3
        stores = {store["ceco"]: store for store in payload["stores"]}
        assert stores["38894"]["applicableActivities"][horno["name"]] is False
        assert stores["38119"]["applicableActivities"][community["name"]] is False

        # Escenario 11: se reproduce el orden exacto del Excel Forms vigente.
        # Si Microsoft inserta o mueve columnas, el registro se arma por encabezado.
        actual = temp / "actual-order.xlsx"
        actual_source = load_workbook(ROOT / "cms" / "Sistema de Evidencias OPS.xlsx", read_only=True)
        actual_headers = [cell.value for cell in next(actual_source["Sheet1"].iter_rows(min_row=1, max_row=1))]
        actual_positions = {str(header).strip(): index for index, header in enumerate(actual_headers) if header}

        def actual_row(values: dict[str, object]) -> list[object]:
            row: list[object] = [""] * len(actual_headers)
            for header, value in values.items():
                row[actual_positions[header]] = value
            return row

        actual_rows = []
        for index, ceco in enumerate(enrique_cecos):
            start, finish = timestamps(50 + index)
            values = {
                "Id": 50 + index,
                "Hora de inicio": start,
                "Hora de finalización": finish,
                "CeCo": ceco,
                ACTIVITY: "Programacion Hornos Merry - Focaccia",
                "¿ Tienes Horno Merry Chef ?": "Sí" if index < 8 else "No",
            }
            if index < 8:
                values["Evidencia_Programacion_Hornos_Merry_Focaccia"] = f"{allowed}/actual-horno-{index}.jpg"
            actual_rows.append(actual_row(values))
        start, finish = timestamps(70)
        actual_rows.append(actual_row({
            "Id": 70, "Hora de inicio": start, "Hora de finalización": finish,
            "CeCo": "38115", ACTIVITY: "Community Board",
            "¿Cuentas con Community Board?": "Sí",
            "Evidencia_Community_Board": f"{allowed}/actual-community.jpg",
        }))
        start, finish = timestamps(71)
        actual_rows.append(actual_row({
            "Id": 71, "Hora de inicio": start, "Hora de finalización": finish,
            "CeCo": "38119", ACTIVITY: "Community Board",
            "¿Cuentas con Community Board?": "No",
        }))
        save_book(actual, actual_headers, actual_rows)
        payload = build_payload(
            actual,
            ROOT / "cms" / "Centro Norte_Directorio.xlsx",
            ROOT / "config" / "settings.json",
            ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx",
        )
        activity_map = {item["name"]: item for item in payload["activities"]}
        assert (activity_map["Programacion Hornos Merry - Focaccia"]["completedStores"], activity_map["Programacion Hornos Merry - Focaccia"]["notApplicableStores"]) == (8, 2)
        assert (activity_map["Community Board"]["completedStores"], activity_map["Community Board"]["notApplicableStores"]) == (1, 1)
        assert payload["quality"]["responseSchema"]["applicabilityIssues"] == {}

        # Escenario 12: respuestas contradictorias se rechazan sin alterar conteos.
        conflicting = temp / "conflicting.xlsx"
        start, finish = timestamps(42)
        save_book(conflicting, conditional_headers, [[
            42, start, finish, "", "Prueba", "38333", "Programacion Hornos Merry - Focaccia",
            "Sí", "No", "", f"{allowed}/conflicto.jpg", "",
        ]])
        rows, schema = load_responses(conflicting, [item["name"] for item in cms_activities])
        assert rows[0]["confirmed"] is False and rows[0]["applicabilityConflict"] is True
        assert schema["applicabilityIssues"]["conflicting-applicability-answers"] == [2]

        # Escenario 13: un archivo renombrado como XLSX se rechaza antes de procesarse.
        damaged = temp / "damaged.xlsx"
        damaged.write_bytes(b"archivo incompleto")
        try:
            load_responses(damaged)
        except ValueError as error:
            assert "dañado" in str(error)
        else:
            raise AssertionError("El XLSX dañado no fue rechazado")

        # Escenario 14: Forms puede quedar sólo con encabezados después de limpiar filas.
        empty = temp / "empty.xlsx"
        save_book(empty, ["CeCo", ACTIVITY, "Evidencia_RollOut"], [])
        payload = build_payload(
            empty,
            ROOT / "cms" / "Centro Norte_Directorio.xlsx",
            ROOT / "config" / "settings.json",
            ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx",
        )
        assert payload["quality"]["responsesRead"] == 0
        assert payload["summary"]["completedCompletions"] == 0
        assert payload["submissions"] == []

    print("Forms dinámico aprobado · orden real · Sí/No implícito · CMS controla visibilidad")


if __name__ == "__main__":
    main()
