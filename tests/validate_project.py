#!/usr/bin/env python3
from __future__ import annotations

import json
import re
import subprocess
import sys
import tempfile
from pathlib import Path
from urllib.parse import unquote, urlsplit

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))
from scripts.build_dashboard import STABILITY_CONTROLS, compact_key, evidence_key, file_sha256, load_responses, safe_evidence_url
from scripts.clean_obsolete import OBSOLETE_FILES

REQUIRED = [
    "index.html", "styles.css", "app.js", "pdf-export.js", "xlsx-export.js", "service-worker.js", "manifest.webmanifest",
    "data/dashboard.json", "exports/Resumen_Evidencias_OPS.xlsx", "exports/Resumen_Evidencias_OPS.pdf", "scripts/build_dashboard.py", "scripts/update_directory.py", "scripts/validate_sources.py", "scripts/clean_obsolete.py", "scripts/export_excel.py", "scripts/export_pdf.py", "scripts/prepare_images.py",
    "scripts/audit_project.py", "config/settings.json", "INSTRUCCION_FORMS.md", "MEJORAS.md",
    "cms/Directorio.xlsx", "cms/Sistema de Evidencias OPS.xlsx",
    "cms/Sistema_Evidencias_OPS_CMS.xlsx", ".github/workflows/build-dashboard.yml", ".nojekyll",
    "assets/icons/icon-64.png", "assets/icons/icon-192.png", "assets/icons/icon-512.png",
    "assets/icons/icon-64.webp", "assets/icons/icon-192.webp", "assets/icons/icon-512.webp", "assets/icons/ops-logo.webp",
    "assets/director/jorge-alcantar.webp", "assets/director/raul-sierra.webp",
    "assets/ui/Damos_Seguimiento.webp", "assets/ui/Un_placer_haber_Ayudado.webp", "tests/build_dynamic_xlsx.js", "tests/build_direct_pdf.js",
    "tests/validate_dynamic_forms_schema.py", "tests/validate_maintenance.py", "scripts/io_utils.py",
]
REQUIRED += [f"assets/dm/{name}.webp" for name in (
    "enrique-cesar", "nancy-carolina", "vanessa-carreno", "veronica-garcia", "yazmin-chabela", "yazmin-garcia"
)]
TEXT_SUFFIXES = {".py", ".js", ".css", ".html", ".md", ".yml", ".yaml", ".json"}


def fail(message: str) -> None:
    raise AssertionError(message)


passed: list[str] = []


def approve(name: str) -> None:
    passed.append(name)


allowed_hosts = {"grupovips-my.sharepoint.com"}
safe_sample = "https://grupovips-my.sharepoint.com/ruta/imagen.jpg#vista"
if safe_evidence_url(safe_sample, allowed_hosts) != safe_sample:
    fail("El enlace SharePoint autorizado fue rechazado")
for unsafe in ("http://grupovips-my.sharepoint.com/imagen.jpg", "https://usuario@grupovips-my.sharepoint.com/imagen.jpg", "https://example.com/imagen.jpg"):
    if safe_evidence_url(unsafe, allowed_hosts):
        fail("La validación aceptó un enlace de evidencia inseguro")


for relative in REQUIRED:
    if not (ROOT / relative).is_file():
        fail(f"Falta archivo requerido: {relative}")
obsolete_present = [relative for relative in OBSOLETE_FILES if (ROOT / relative).exists()]
if obsolete_present:
    fail("Persisten archivos obsoletos: " + ", ".join(obsolete_present))
public_docs = "\n".join((ROOT / name).read_text(encoding="utf-8") for name in ("ARCHIVOS.md", "MEJORAS.md"))
obsolete_references = [relative for relative in OBSOLETE_FILES if relative in public_docs]
if obsolete_references:
    fail("La documentación conserva rutas obsoletas: " + ", ".join(obsolete_references))
mojibake_codepoints = {0x00C2, 0x00C3, 0x00E2}
encoding_issues = []
for source in ROOT.rglob("*"):
    if not source.is_file() or ".git" in source.parts or source.suffix.casefold() not in TEXT_SUFFIXES:
        continue
    try:
        source_text = source.read_text(encoding="utf-8")
    except UnicodeDecodeError:
        encoding_issues.append(source.relative_to(ROOT).as_posix())
        continue
    if any(ord(character) in mojibake_codepoints for character in source_text):
        encoding_issues.append(source.relative_to(ROOT).as_posix())
if encoding_issues:
    fail("Archivos con codificación dañada: " + ", ".join(sorted(encoding_issues)))
approve("01 · Archivos requeridos y limpieza de obsoletos")

data = json.loads((ROOT / "data/dashboard.json").read_text(encoding="utf-8"))
manifest = json.loads((ROOT / "manifest.webmanifest").read_text(encoding="utf-8"))
html = (ROOT / "index.html").read_text(encoding="utf-8")
css = (ROOT / "styles.css").read_text(encoding="utf-8")
js = (ROOT / "app.js").read_text(encoding="utf-8")
sw = (ROOT / "service-worker.js").read_text(encoding="utf-8")
workflow = (ROOT / ".github/workflows/build-dashboard.yml").read_text(encoding="utf-8")

if data.get("schemaVersion") != 12:
    fail("Versión del contrato JSON incorrecta")
if data.get("project") != "Sistema de Evidencias OPS" or data.get("region") != "Todas las regiones":
    fail("Identidad del proyecto incorrecta")
if not re.fullmatch(r"[0-9a-f]{16}", data.get("buildVersion", "")):
    fail("La versión Python para invalidar caché es incorrecta")
if data.get("sources", {}).get("directorySheet") != "Directorio":
    fail("No se utilizó la hoja configurada del directorio")
if data.get("sources", {}).get("cms") != "Sistema_Evidencias_OPS_CMS.xlsx":
    fail("Python no está leyendo el Excel CMS")
if not re.fullmatch(r"\d{2}/\d{2}/\d{4} \d{2}:\d{2}", data.get("lastUpdatedDisplay", "")):
    fail("Última actualización incorrecta")
summary = data.get("summary", {})
if not data.get("dms") or not data.get("stores") or not data.get("activities"):
    fail("El dashboard quedó sin alcance operativo")
if summary.get("dms") != len(data["dms"]) or summary.get("stores") != len(data["stores"]) or summary.get("activities") != len(data["activities"]):
    fail("Los conteos no coinciden con el alcance generado")
if data.get("calendar", {}).get("active") != len(data["activities"]):
    fail("Las actividades vigentes del CMS no fueron calculadas")
for source_key, path in (
    ("responsesSha256", ROOT / "cms" / "Sistema de Evidencias OPS.xlsx"),
    ("directorySha256", ROOT / "cms" / "Directorio.xlsx"),
    ("cmsSha256", ROOT / "cms" / "Sistema_Evidencias_OPS_CMS.xlsx"),
):
    if data.get("sources", {}).get(source_key) != file_sha256(path):
        fail(f"La huella de la fuente {source_key} no coincide")

sample = next((store for store in data.get("stores", []) if store.get("ceco") == "38115"), None)
if not sample or sample.get("store") != "Zona Azul" or sample.get("dm") != "Yazmin Haydee Garcia Gonzalez":
    fail("Falló el cruce 38115 → Zona Azul → Yazmin Haydee")
activity_names = [item["name"] for item in data.get("activities", [])]
calculated_exclusions = 0
for store in data.get("stores", []):
    applicability = store.get("applicableActivities", {})
    expected = sum(applicability.get(name, True) is not False for name in activity_names)
    excluded = len(activity_names) - expected
    calculated_exclusions += excluded
    if store.get("expected") != expected or store.get("notApplicable") != excluded:
        fail(f"Aplicabilidad inconsistente para CeCo {store.get('ceco')}")
    if any(store.get("activities", {}).get(name) and applicability.get(name, True) is False for name in activity_names):
        fail(f"CeCo {store.get('ceco')} contabiliza una actividad excluida")
if summary.get("notApplicableCompletions") != calculated_exclusions:
    fail("La resta implícita de actividades no coincide con las respuestas Sí/No")
if len(data.get("regions", [])) != 4 or summary.get("regions") != 4 or summary.get("stores") != 357:
    fail("El alcance multirregión del Directorio no quedó publicado")
directory_status = data.get("sources", {}).get("directoryStatus", {})
if directory_status.get("includedStatuses") != ["Abierta"] or directory_status.get("includedStores") != 357 or directory_status.get("excludedStores") != 15:
    fail("El CMS no controla de forma auditable las tiendas abiertas")
if any(store.get("status") != "Abierta" for store in data.get("stores", [])):
    fail("Una tienda no abierta entró en los conteos del dashboard")
if sum(item.get("photoStatus") == "Disponible" for item in data.get("dms", [])) != 6:
    fail("Las seis fotografías existentes no quedaron vinculadas")
if not any(item.get("photoStatus") == "Pendiente" for item in data.get("dms", [])):
    fail("Los DM nuevos no quedaron marcados con foto pendiente")
if data.get("quality", {}).get("unknownCeCos") or data.get("quality", {}).get("unsafeEvidenceRows"):
    fail("Calidad inicial incorrecta")
if any("email" in row or "submittedBy" in row for row in data.get("submissions", [])):
    fail("El JSON público expone correo o respondente")
published = [row for row in data.get("submissions", []) if row.get("valid")]
forms_responses, forms_schema = load_responses(
    ROOT / "cms" / "Sistema de Evidencias OPS.xlsx",
    [item["name"] for item in data.get("activities", [])],
)
active_by_key = {compact_key(item["name"]): item["name"] for item in data.get("activities", [])}
stores_by_ceco = {item["ceco"]: item for item in data.get("stores", [])}
latest_excel_by_pair = {}
for row in forms_responses:
    activity = active_by_key.get(compact_key(row["activity"]))
    evidence_url = safe_evidence_url(row["evidence"], allowed_hosts)
    if not activity or row["ceco"] not in stores_by_ceco or not row["confirmed"] or row["explicitNo"] or not evidence_url:
        continue
    pair = (row["ceco"], activity)
    row_sort = (row["finished"].isoformat() if row["finished"] else "", row["row"])
    current = latest_excel_by_pair.get(pair)
    if current is None or row_sort > current[0]:
        latest_excel_by_pair[pair] = (row_sort, evidence_url)
expected_excel_links = {pair: item[1] for pair, item in latest_excel_by_pair.items()}
published_excel_links = {(row["ceco"], row["activity"]): row["evidenceUrl"] for row in published}
if data.get("quality", {}).get("evidenceLinksPublished") != len(published) or summary.get("validResponses") != len(published):
    fail("El conteo dinámico de vínculos publicados no coincide con las respuestas válidas")
if any(not row.get("evidenceFileName") or not row.get("evidenceUrl") or row.get("evidenceLinkLabel") != f"Link_{row.get('evidenceKey')}" or urlsplit(row["evidenceUrl"]).hostname not in allowed_hosts for row in published):
    fail("Nombre de archivo o vínculo directo inválido")
if published_excel_links != expected_excel_links:
    missing = len(set(expected_excel_links).difference(published_excel_links))
    unexpected = len(set(published_excel_links).difference(expected_excel_links))
    changed = sum(published_excel_links.get(pair) != url for pair, url in expected_excel_links.items() if pair in published_excel_links)
    fail(f"La última evidencia por tienda y actividad no coincide: faltan {missing}, sobran {unexpected}, cambiaron {changed}")
if not forms_schema["evidenceHeaders"] or forms_schema["rowConflicts"] or forms_schema["evidenceIssues"] or forms_schema.get("applicabilityIssues"):
    fail("El esquema dinámico de evidencias no fue detectado correctamente")
evidence_header_matches = forms_schema.get("evidenceHeaderMatch", {})
evidence_header_map = forms_schema.get("evidenceHeaderMap", {})
if any(match not in {"exact", "affinity", "generic", "unverified"} for match in evidence_header_matches.values()):
    fail("Un encabezado de evidencia produjo una relación ambigua o insegura")
for header, match in evidence_header_matches.items():
    mapped_key = evidence_header_map.get(header)
    if match in {"exact", "affinity"} and mapped_key not in active_by_key:
        fail(f"El encabezado {header} declara una actividad CMS inexistente")
    if match == "unverified" and mapped_key in active_by_key:
        fail(f"El encabezado {header} dejó sin relacionar una actividad activa del CMS")
for row in published:
    expected_name = unquote(urlsplit(row["evidenceUrl"]).path.rsplit("/", 1)[-1])
    if row["evidenceFileName"] != expected_name:
        fail("El nombre de archivo no coincide con el vínculo del Excel")
if data.get("submissions") and data["submissions"][0].get("timestampDisplay") != data.get("lastUpdatedDisplay"):
    fail("La última actualización no coincide con la respuesta más reciente")
for submission in published:
    store = stores_by_ceco.get(submission["ceco"])
    if not store or store["store"] != submission["store"] or store["dm"] != submission["dm"]:
        fail("Una respuesta no cruzó correctamente contra el directorio")
    if store.get("activities", {}).get(submission["activity"]) is not True:
        fail("Una respuesta válida no quedó contabilizada por nombre de actividad")
    if submission["evidenceKey"] != evidence_key(submission["activity"], submission["ceco"]):
        fail("La llave de evidencia no se construyó desde actividad y CeCo")
approve("02 · CMS, conteos, CeCo y evidencias seguras")

with tempfile.TemporaryDirectory() as temp_dir:
    generated = Path(temp_dir) / "dashboard.json"
    subprocess.run([sys.executable, str(ROOT / "scripts/build_dashboard.py"), "--output", str(generated)], cwd=ROOT, check=True, stdout=subprocess.DEVNULL)
    fresh = json.loads(generated.read_text(encoding="utf-8"))
for payload in (data, fresh):
    payload.pop("generatedAt", None)
if data != fresh:
    fail("data/dashboard.json está desincronizado")
approve("03 · Python sincronizado con la última actualización")

static_excel = load_workbook(ROOT / "exports" / "Resumen_Evidencias_OPS.xlsx", data_only=False)
if static_excel.sheetnames != ["Resumen", "Tiendas", "Actividades"]:
    fail("El Excel Python no contiene las tres vistas ejecutivas")
if any(not str(static_excel[sheet]["A1"].fill.fgColor.rgb).endswith("002E24") for sheet in static_excel.sheetnames):
    fail("Los títulos del Excel Python no conservan el verde oscuro")
expected_summary_formula = "=IFERROR(A6/(A6+C6),0)"
if static_excel["Resumen"]["E6"].value != expected_summary_formula or static_excel["Resumen"]["A6"].number_format != "#,##0" or static_excel["Resumen"]["E6"].number_format != "0.0%" or static_excel["Resumen"]._charts:
    fail("El resumen Excel no conserva fórmula, formato numérico o limpieza visual")
for sheet_name, header_row in (("Resumen", 9), ("Tiendas", 4), ("Actividades", 4)):
    headers = [cell.value for cell in static_excel[sheet_name][header_row]]
    if "Pendientes" not in headers or "Decisión" not in headers or any(label in headers for label in ("Aplican", "No aplica", "N/A")):
        fail(f"La hoja {sheet_name} no está enfocada únicamente en Realizadas y Pendientes")
with tempfile.TemporaryDirectory() as temp_dir:
    dynamic_excel = Path(temp_dir) / "dinamico.xlsx"
    subprocess.run(["node", str(ROOT / "tests" / "build_dynamic_xlsx.js"), str(dynamic_excel)], cwd=ROOT, check=True, stdout=subprocess.DEVNULL)
    dynamic_book = load_workbook(dynamic_excel, data_only=False)
    dm_sheet = dynamic_book["Tiendas"]
    if dynamic_book.sheetnames != ["Resumen", "Tiendas", "Actividades"] or dynamic_book["Resumen"]["B5"].value != 0.014 or dynamic_book["Resumen"]["B5"].number_format != "0.0%":
        fail("El motor XLSX dinámico generó un libro inválido")
    if any(dynamic_book[sheet]["A1"].fill.fgColor.rgb != "FF002E24" for sheet in dynamic_book.sheetnames):
        fail("El título verde oscuro no se aplicó a todas las pestañas dinámicas")
    if [cell.value for cell in dm_sheet[4]] != ["CeCo", "Tienda", "Roll Out", "Rack FHW", "QR - Qualtrics", "Mandil Verde", "Realizadas", "Pendientes", "% Avance", "Estado", "Decisión"]:
        fail("La hoja Tiendas no contiene el detalle por actividad")
    if dm_sheet["G5"].value != "=SUM(C5:F5)" or dm_sheet["H5"].value != "=COUNT(C5:F5)-SUM(C5:F5)" or dm_sheet["I5"].value != "=IFERROR(SUM(C5:F5)/COUNT(C5:F5),0)" or dm_sheet["I5"].number_format != "0.0%":
        fail("Realizadas, Pendientes o porcentaje del DM no son auditables")
    if dm_sheet["A1"].fill.fgColor.rgb != "FF002E24" or dm_sheet["C5"].fill.fgColor.rgb != "FF1E3932" or dm_sheet["F5"].fill.fgColor.rgb != "FFE9F4EF" or dm_sheet["J5"].fill.fgColor.rgb != "FFFFF0D5" or dm_sheet["K5"].value != "Dar seguimiento" or dm_sheet["D5"].value not in (None, ""):
        fail("El contraste del título o los estados realizados/pendientes no es consistente")
approve("04 · XLSX regional y dinámico con formatos congruentes")
with tempfile.TemporaryDirectory() as temp_dir:
    direct_pdf = Path(temp_dir) / "directo.pdf"
    subprocess.run(["node", str(ROOT / "tests" / "build_direct_pdf.js"), str(direct_pdf)], cwd=ROOT, check=True, stdout=subprocess.DEVNULL)
    pdf_bytes = direct_pdf.read_bytes()
    if not pdf_bytes.startswith(b"%PDF-1.4") or not pdf_bytes.rstrip().endswith(b"%%EOF"):
        fail("El motor PDF directo generó un archivo inválido")
regional_pdf = (ROOT / "exports" / "Resumen_Evidencias_OPS.pdf").read_bytes()
if not regional_pdf.startswith(b"%PDF-") or len(regional_pdf) < 20_000:
    fail("El PDF regional Python no fue generado correctamente")
approve("05 · PDF regional Python y descarga directa válidos")

for text in ["Sistema de Evidencia OPS", "Dashboard de Avance de Actividades", "Resumen", "Organigrama", "Región | Centro's", "Ranking DM", "Actividades", "Tiendas", "Evidencias", "Actividad", "Tienda", "Link del archivo", "filter-region", "evidence-details", "evidence-filter-region", "evidence-filter-dm", "evidence-filter-activity", "evidence-filter-store", "export-image", "export-pdf", "export-excel", "export-modal", "Damos_Seguimiento.webp", "activity-focus-table", "evidence-grid", "dm-team", "store-table", "Director Starbucks México", "Raúl Sierra", "Starbucks México · Operaciones"]:
    if text not in html:
        fail(f"Interfaz simplificada incompleta: {text}")
nav_order = [html.index(f'href="#{item}"') for item in ("resumen", "ranking", "actividades", "tiendas", "evidencias")]
section_order = [html.index(f'id="{item}"') for item in ("resumen", "ranking", "actividades", "tiendas", "evidencias")]
if nav_order != sorted(nav_order) or section_order != sorted(section_order):
    fail("Orden de navegación o secciones incorrecto")
if "Última hora del dato actualizado" in html or re.search(r'<details[^>]+id="evidence-details"[^>]+open', html):
    fail("Fecha de corte o panel de soporte no respetan el diseño solicitado")
store_renderer = js[js.index("function renderStores"):js.index("function syncFilterUrl")]
if "<th>DM</th>" in html or "esc(store.dm)" in store_renderer or 'colspan="7"' in store_renderer:
    fail("La tabla Tiendas todavía muestra la columna DM")
for forbidden in ["class=\"sidebar\"", "side-nav", "data-route=", "routeTo(", "--sidebar", "guide-steps", "priority-stores", "quality-strip", "Atención prioritaria", "De mayor a menor avance", "Detalle dinámico", "id=\"filter-notice\"", "id=\"activity-context\"", "id=\"evidence-title\"", "id=\"team-title\"", "id=\"stores-title\"", "id=\"store-summary\"", "id=\"active-scope\"", "id=\"toggle-dates\"", "id=\"commitment-dates\"", "renderActiveScope"]:
    if forbidden in html + js + css:
        fail(f"Elemento lateral obsoleto aún presente: {forbidden}")
approve("06 · Navegación lineal y sin bloques obsoletos")
stability_controls = data.get("quality", {}).get("stabilityControls", {})
if tuple(stability_controls) != STABILITY_CONTROLS or not all(stability_controls.values()) or data.get("quality", {}).get("stabilityScore") != "10/10":
    fail("Los 10 controles Python de estabilidad no están activos")
for required in [".activity-table-shell { overflow-x: clip", ".activity-focus-table { width: 100%; min-width: 0; table-layout: fixed", ".activity-focus-table { display: table", ".activity-focus-table .activity-focus-row { display: table-row", ".activity-focus-table .activity-focus-row td { display: table-cell"]:
    if required not in css:
        fail(f"Actividades no está adaptada a móvil: {required}")
if re.search(r"\.activity-focus-table\s*\{[^}]*min-width:\s*(?:8\d\d|9\d\d|\d{4,})px", css):
    fail("Actividades conserva un ancho mínimo que provoca desplazamiento horizontal")
approve("06B · Actividades en una fila y sin desplazamiento horizontal en móvil")
for text in ["renderSummary", "renderActivities", "renderEvidence", "populateEvidenceFilters", "evidenceFilters", "evidenceLinkLabel", "exportRows", "renderTeam", "renderStores", "syncFilterUrl", "clearDashboardFilters", "back-to-top", "beginExport", "finishExport", "exportImage", "exportPdf", "exportExcel", "buildExcelSpec", "renderPdfPages", "exportProfile", "exportActivityLabel", "exportAdvanceLabel", "AVANCE REGIÓN", "icon-192.webp", "spreadsheetColumn", "Detalle de actividades por tienda", "1 = Realizada · 0 = Pendiente", "acceptExportConfirmation", "Aceptar y descargar", "Valida tu archivo", "Carpeta Descargas", "Cerrar exportación", "export-close", "URL.revokeObjectURL", "AVANCE REALIZADO", "PENDIENTES", "% AVANCE", "Un_placer_haber_Ayudado.webp", "noopener noreferrer", "referrerpolicy", "serviceWorker", "deadlineLabel", "focusRank"]:
    if text not in js:
        if text not in html + css:
            fail(f"Funcionalidad faltante: {text}")
for forbidden in ("export-modal-open", "Abrir PDF", "Ver imagen", "Descargar Excel", ">Ver archivo<"):
    if forbidden in html + js + css:
        fail(f"La confirmación final conserva una acción obsoleta: {forbidden}")
if "tiendas · ${dm.completed} realizadas" in js:
    fail("Ranking DM todavía muestra realizadas junto a tiendas")
for obsolete_summary in ('id="filter-summary"', "renderFilterSummary", "data-clear-dashboard-filters"):
    if obsolete_summary in html + js + css:
        fail(f"Resumen redundante todavía visible: {obsolete_summary}")
for redundant_export_text in (
    'fillText(meta.motto, 800, 45)',
    'fillText(meta.motto, 800, 55)',
    '`Actividad · ${exportActivityLabel()}`',
    '`ACTIVIDAD  ${exportActivityLabel()}`',
    '`${director.role} · ${director.name}`',
):
    if redundant_export_text in js:
        fail(f"Exportación redundante: {redundant_export_text}")
if js.count("./assets/icons/icon-192.webp") < 2 or js.count("context.fillRect(1320,") < 2:
    fail("PDF e imagen no comparten icono grande o recuadro de avance")
excel_export_source = js[js.index("function buildExcelSpec"):js.index("async function exportExcel")]
for required_excel_context in ("const activityLabel = exportActivityLabel()", "exportAdvanceLabel()", "${scope} · ${activityLabel}"):
    if required_excel_context not in excel_export_source:
        fail(f"Excel perdió el filtro dinámico: {required_excel_context}")
if "event.target === event.currentTarget" in js or "URL.revokeObjectURL(state.exportUrl)" not in js or "link.download = exportInfo.filename" not in js:
    fail("La descarga automática, el cierre explícito o la liberación de memoria están incompletos")
approve("07 · Filtros, confirmación y exportaciones del alcance actual")
for cache_behavior in ("enforceBuildVersion", "BUILD_STORAGE_KEY", "localStorage", "sessionStorage", "window.location.replace", 'headers: { "Cache-Control": "no-cache" }', "loadScriptOnce", "loadExportEngine"):
    if cache_behavior not in js:
        fail(f"Actualización automática sin caché incompleta: {cache_behavior}")
for cache_control in ("sistema-evidencias-ops-v23", "staleWhileRevalidate", 'cache: "no-store"', "skipWaiting", "clients.claim", "CACHE_PREFIX", "CLEAR_ALL_CACHES"):
    if cache_control not in sw:
        fail(f"Actualización PWA incompleta: {cache_control}")
if "Sistema_Evidencias_OPS_CMS.xlsx" in sw:
    fail("El Excel CMS no debe publicarse en la caché web")
core_cache = sw[sw.index("const CORE"):sw.index("];", sw.index("const CORE"))]
if any(path in core_cache for path in ("/exports/", "/assets/dm/", "/assets/ui/", "icon-192")):
    fail("La instalación PWA todavía precarga archivos pesados no esenciales")
if 'src="./pdf-export.js"' in html or 'src="./xlsx-export.js"' in html or "Date.now()" in js[js.index("async function loadData"):js.index("async function refreshApplicationData")]:
    fail("La carga inicial conserva motores pesados o genera entradas de caché únicas")
gitignore_path = ROOT / ".gitignore"
if gitignore_path.is_file():
    gitignore = gitignore_path.read_text(encoding="utf-8")
    for ignored in ("__pycache__/", "*.py[cod]", "*.tmp", "cms/~$*.xlsx"):
        if ignored not in gitignore:
            fail(f"La limpieza local no ignora {ignored}")
if "window.print" in js or "Tiendas realizadas" in js:
    fail("La descarga directa o el KPI inicial aún conserva comportamiento obsoleto")
if "Todas las actividades · Ranking regional de mayor a menor avance" in js + (ROOT / "scripts/export_pdf.py").read_text(encoding="utf-8"):
    fail("El PDF aún conserva el subtítulo regional eliminado")
export_pdf_source = (ROOT / "scripts/export_pdf.py").read_text(encoding="utf-8")
if "% PENDIENTE" in js + export_pdf_source or "PÁGINA ${pageIndex" in js or "Página {page_index" in export_pdf_source:
    fail("La exportación conserva porcentaje pendiente o numeración de página")
excel_spec_source = js[js.index("function buildExcelSpec"):js.index("async function exportExcel")]
for forbidden_export_label in ('"Aplican"', '"No aplica"', '"N/A"', "NO APLICA", "REALIZADAS / APLICAN"):
    if forbidden_export_label in excel_spec_source + export_pdf_source:
        fail(f"La exportación todavía muestra {forbidden_export_label}")
if '"Pendientes"' not in excel_spec_source or "AVANCE REALIZADO" not in js + export_pdf_source:
    fail("Las exportaciones no están enfocadas en avance realizado y pendientes")
for required in ("SUM(${activityRange})", "COUNT(${activityRange})-SUM(${activityRange})", "profile.photo", 'role: "DM"'):
    if required not in js:
        fail(f"Detalle DM incompleto en exportaciones: {required}")
if "guide" in data:
    fail("La guía eliminada todavía se publica en el JSON")
approve("08 · PWA, descarga directa y mensaje final simplificados")
if [item.get("rank") for item in data.get("dms", [])] != list(range(1, len(data.get("dms", [])) + 1)):
    fail("Ranking DM inválido")
director = data.get("report", {}).get("regionalDirector", {})
organization = data.get("organization", {})
if data.get("report", {}).get("motto") != "JUNTÉMONOS MÁS" or data.get("report", {}).get("footerLabel") != "Starbucks México · Operaciones" or director.get("name") != "Jorge Alcantar" or director.get("role") != "Director Regional" or organization.get("nationalDirector", {}).get("name") != "Raúl Sierra" or len(organization.get("regionalDirectors", [])) != 4 or any(not {"commitmentDateDisplay", "deadlineLabel", "deadlineTone", "focusRank"}.issubset(item) for item in data.get("activities", [])):
    fail("Exportación o fechas compromiso no fueron preparadas por Python")
focus = data.get("activities", [])
if [item.get("focusRank") for item in focus] != list(range(1, len(focus) + 1)):
    fail("El foco de actividades no es consecutivo")
if '"Aplican"' in js[js.index("function renderSummary"):js.index("function renderActivities")] or "aplican${" in js[js.index("function renderTeam"):js.index("function renderStores")]:
    fail("El resumen o las tarjetas DM conservan la palabra Aplican")
if not any(icon.get("sizes") == "64x64" for icon in manifest.get("icons", [])):
    fail("El nuevo logo no está configurado en todos los tamaños")
approve("09 · Ranking, fotografía DM e identidad ejecutiva")
for text in ["pip check", "python -X utf8 scripts/safe_maintenance.py --force", "python -X utf8 scripts/clean_obsolete.py --check", "git add -- data/dashboard.json exports/Resumen_Evidencias_OPS.xlsx exports/Resumen_Evidencias_OPS.pdf"]:
    if text not in workflow:
        fail(f"Workflow incompleto: {text}")
for text in ["PYTHONUTF8: '1'", "PYTHONPYCACHEPREFIX: /tmp/evidencias-ops-pycache", "node --check service-worker.js", "git diff --check", "set -euo pipefail", "git diff --cached --quiet", "git ls-files --error-unmatch", 'obsolete_test="tests/validate_horno_applicability.py"']:
    if text not in workflow:
        fail(f"Publicación no idempotente: falta {text}")
if "git add -A -- tests/validate_horno_applicability.py" in workflow:
    fail("El workflow conserva el pathspec directo que falla si el archivo no existe")
approve("10 · Workflow completo: limpiar, generar, validar y publicar")

if len(passed) != 11:
    fail(f"Se esperaban 11 validaciones y se ejecutaron {len(passed)}")
print("Validación aprobada · 11/11 controles")
for check in passed:
    print(f"OK {check}")
print("CMS Excel → Python → un JSON consolidado")
print(f"{summary['stores']} tiendas · {summary['activities']} actividades vigentes · {summary['dms']} DM + 1 Director Regional")
if published:
    sample_submission = published[0]
    print(f"{sample_submission['evidenceKey']} → {sample_submission['store']} · vínculo SharePoint validado")
else:
    print("Forms sin respuestas válidas · tablero vacío aceptado")
print("Imagen/PDF: Todos los DM → ranking DM · Un DM → tiendas descendentes")
print("Excel: resumen rápido + detalle + actividades")
